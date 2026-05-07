"""
AI Smart Attendance System - MVP (Data-Driven)
A fully functional Streamlit application for attendance management.
Uses Streamlit session_state for real data management.
Cloud-deployment ready: works via browser link on any device.
"""

import streamlit as st
from datetime import datetime, timedelta
import json
import base64
import os
import pandas as pd
import threading
import time
from io import BytesIO
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Bootstrap: ensure required dirs/files exist (critical for cloud) ──
try:
    from init_data import ensure_data_files
    ensure_data_files()
except Exception:
    Path("data").mkdir(exist_ok=True)
    Path("dataset").mkdir(exist_ok=True)
    if not Path("data.json").exists():
        import json as _j
        Path("data.json").write_text(_j.dumps({"students":[],"timetable":{},"period_attendance":[]}, indent=2))
    if not Path("data/erp_attendance.json").exists():
        Path("data/erp_attendance.json").write_text("{}")


# ============================================================================
# A. PAGE CONFIGURATION & STYLING
# ============================================================================

st.set_page_config(
    page_title="AttendAI — Next-Gen AI Attendance Portal",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for state-of-the-art premium futuristic UI/UX (AttendAI Cyberpunk Theme)
st.markdown("""
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Syne:wght@700;800&family=JetBrains+Mono:wght@400;500;700&display=swap" rel="stylesheet">
    
    <style>
        /* Global Typography Override (Syne for Headers, Inter for Body, JetBrains Mono for Metrics) */
        html, body, [class*="css"], .stMarkdown, p, span, label, select, input, button {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
        }
        
        h1, h2, h3, h4, h5, h6, .syne-title {
            font-family: 'Syne', -apple-system, sans-serif !important;
            letter-spacing: -0.03em !important;
            font-weight: 800 !important;
            background: linear-gradient(135deg, #ffffff 30%, #a78bfa 100%) !important;
            -webkit-background-clip: text !important;
            -webkit-text-fill-color: transparent !important;
        }
        
        /* Futuristic Cyberpunk SaaS Background with Ambient Glowing Orbs */
        .stAppViewContainer {
            background-color: #060912 !important;
            background-image: 
                radial-gradient(at 0% 0%, rgba(79, 158, 255, 0.08) 0px, transparent 40%),
                radial-gradient(at 95% 10%, rgba(167, 139, 250, 0.08) 0px, transparent 40%),
                radial-gradient(at 10% 90%, rgba(52, 211, 153, 0.06) 0px, transparent 40%),
                radial-gradient(at 90% 85%, rgba(244, 114, 182, 0.06) 0px, transparent 40%) !important;
            background-attachment: fixed !important;
            color: #cbd5e1 !important;
        }
        
        header[data-testid="stHeader"] {
            display: none !important;
        }
        
        .main .block-container {
            max-width: 95% !important;
            padding-top: 1.5rem !important;
            padding-left: 3rem !important;
            padding-right: 3rem !important;
        }
        
        /* Sleek Modern Scrollbars */
        ::-webkit-scrollbar {
            width: 6px;
            height: 6px;
        }
        ::-webkit-scrollbar-track {
            background: #060912;
        }
        ::-webkit-scrollbar-thumb {
            background: rgba(167, 139, 250, 0.3);
            border-radius: 10px;
        }
        ::-webkit-scrollbar-thumb:hover {
            background: rgba(79, 158, 255, 0.6);
        }
        
        /* Glassmorphic Frosted Cards (inspired by Linear & Vercel) */
        .animated-card, div[data-testid="stForm"] {
            background: rgba(12, 17, 33, 0.6) !important;
            backdrop-filter: blur(24px) !important;
            -webkit-backdrop-filter: blur(24px) !important;
            border: 1px solid rgba(255, 255, 255, 0.06) !important;
            border-radius: 20px !important;
            padding: 24px !important;
            margin-bottom: 20px !important;
            box-shadow: 0 15px 35px -10px rgba(0, 0, 0, 0.5) !important;
            transition: all 0.4s cubic-bezier(0.16, 1, 0.3, 1) !important;
        }
        
        .animated-card:hover, div[data-testid="stForm"]:hover {
            transform: translateY(-4px) !important;
            border-color: rgba(167, 139, 250, 0.3) !important;
            box-shadow: 
                0 30px 60px -15px rgba(0, 0, 0, 0.6), 
                0 0 20px -2px rgba(167, 139, 250, 0.15) !important;
        }
        
        /* Laser Scanner Scanning Line Animation */
        .scanner-container {
            position: relative;
            overflow: hidden;
            border-radius: 16px;
        }
        
        .scanner-line {
            position: absolute;
            width: 100%;
            height: 3px;
            background: linear-gradient(90deg, transparent, #4f9eff, #a78bfa, transparent);
            box-shadow: 0 0 8px #4f9eff, 0 0 15px #a78bfa;
            animation: scanAnimation 2s infinite linear;
            z-index: 10;
        }
        
        @keyframes scanAnimation {
            0% { top: 0%; }
            50% { top: 100%; }
            100% { top: 0%; }
        }
        
        /* Active Class Pulsing Border Overlay */
        .active-period-card {
            border-color: rgba(52, 211, 153, 0.35) !important;
            background: linear-gradient(135deg, rgba(52, 211, 153, 0.05) 0%, rgba(12, 17, 33, 0.8) 100%) !important;
            animation: cardPulse 3s infinite alternate ease-in-out;
        }
        
        @keyframes cardPulse {
            0% { box-shadow: 0 10px 30px -10px rgba(52, 211, 153, 0.1); }
            100% { box-shadow: 0 10px 35px -5px rgba(52, 211, 153, 0.25); }
        }
        
        /* Hide Streamlit Native Sidebar & Controls completely */
        [data-testid="stSidebar"], section[data-testid="stSidebar"], [data-testid="collapsedControl"] {
            display: none !important;
        }
        
        /* ============================================================
           BULLETPROOF WIDGET THEMING - All Streamlit Form Elements
           ============================================================ */

        /* Widget labels */
        label[data-testid="stWidgetLabel"], label[data-testid="stWidgetLabel"] * {
            color: #94a3b8 !important;
            font-weight: 600 !important;
            font-size: 13px !important;
            text-transform: uppercase !important;
            letter-spacing: 0.06em !important;
        }

        /* --- SELECTBOX: Outer container & inner control bar --- */
        div[data-baseweb="select"] {
            background: rgba(6, 9, 18, 0.9) !important;
            border-radius: 12px !important;
        }
        /* The actual visible select bar (control) */
        div[data-baseweb="select"] > div,
        div[data-baseweb="select"] div[class*="control"],
        div[data-baseweb="select"] div[class*="ValueContainer"],
        div[data-baseweb="select"] div[class*="container"] {
            background: rgba(6, 9, 18, 0.9) !important;
            background-color: rgba(6, 9, 18, 0.9) !important;
            border: 1px solid rgba(79, 158, 255, 0.2) !important;
            border-radius: 12px !important;
            color: #ffffff !important;
        }
        /* Selected value text inside selectbox */
        div[data-baseweb="select"] span,
        div[data-baseweb="select"] div[class*="singleValue"],
        div[data-baseweb="select"] div[class*="placeholder"] {
            color: #ffffff !important;
        }
        /* Dropdown arrow icon color */
        div[data-baseweb="select"] svg {
            fill: #94a3b8 !important;
        }
        /* Dropdown list popup */
        div[data-baseweb="select"] ul,
        div[role="listbox"],
        div[data-baseweb="menu"],
        div[data-baseweb="menu"] ul,
        div[data-baseweb="menu"] li {
            background-color: #0c1121 !important;
            color: #ffffff !important;
            border: 1px solid rgba(79, 158, 255, 0.15) !important;
            border-radius: 12px !important;
        }
        div[role="option"], div[role="option"] * {
            color: #ffffff !important;
            background-color: transparent !important;
        }
        div[role="option"]:hover, div[role="option"][aria-selected="true"] {
            background-color: rgba(79, 158, 255, 0.15) !important;
            color: #4f9eff !important;
        }

        /* --- TEXT INPUT, NUMBER INPUT, TEXT AREA --- */
        div[data-baseweb="input"],
        div[data-baseweb="input"] > div,
        div[data-baseweb="textarea"],
        div[data-baseweb="textarea"] > div {
            background: rgba(6, 9, 18, 0.9) !important;
            background-color: rgba(6, 9, 18, 0.9) !important;
            border: 1px solid rgba(79, 158, 255, 0.2) !important;
            border-radius: 12px !important;
        }
        div[data-baseweb="input"] input,
        div[data-baseweb="textarea"] textarea,
        input[type="text"], input[type="number"], input[type="date"], textarea {
            background: transparent !important;
            background-color: transparent !important;
            color: #ffffff !important;
            caret-color: #4f9eff !important;
        }
        div[data-baseweb="input"]:focus-within,
        div[data-baseweb="textarea"]:focus-within {
            border-color: rgba(79, 158, 255, 0.55) !important;
            box-shadow: 0 0 0 3px rgba(79, 158, 255, 0.1) !important;
        }

        /* --- RADIO BUTTONS --- */
        div[data-testid="stRadio"] label,
        div[data-testid="stRadio"] label *  {
            color: #cbd5e1 !important;
        }

        /* --- DATE INPUT CALENDAR POPOVER --- */
        div[data-baseweb="calendar"],
        div[data-baseweb="calendar"] *,
        div[data-baseweb="popover"],
        div[data-baseweb="popover"] * {
            background-color: #0c1121 !important;
            color: #ffffff !important;
            border-color: rgba(79, 158, 255, 0.2) !important;
        }
        div[data-baseweb="calendar"] button {
            color: #ffffff !important;
            background: transparent !important;
        }
        div[data-baseweb="calendar"] button:hover {
            background: rgba(79, 158, 255, 0.2) !important;
            color: #4f9eff !important;
        }

        /* --- PLACEHOLDERS --- */
        ::placeholder, input::placeholder, textarea::placeholder {
            color: #64748b !important;
            opacity: 1 !important;
        }

        /* --- DATAFRAME / TABLE CELLS --- */
        div[data-testid="stDataFrame"] iframe,
        div[data-testid="stDataFrame"] * {
            color: #ffffff !important;
        }
        /* Force dataframe header & cell backgrounds */
        div[data-testid="stDataFrame"] [class*="dvn-scroller"] {
            background: rgba(6, 9, 18, 0.85) !important;
        }

        /* --- GENERAL GLOBAL TEXT fallback --- */
        .stMarkdown p, .stMarkdown li, .stMarkdown span,
        div[data-testid="stMarkdownContainer"] p,
        div[data-testid="stMarkdownContainer"] span {
            color: #cbd5e1 !important;
        }

        /* --- CHECKBOX --- */
        div[data-testid="stCheckbox"] label,
        div[data-testid="stCheckbox"] span {
            color: #cbd5e1 !important;
        }

        /* --- CAPTION & INFO text --- */
        div[data-testid="stCaptionContainer"] p,
        small, .stCaption {
            color: #64748b !important;
        }
        
        /* Neural Metric Customization (Monospace Metrics) */
        div[data-testid="stMetric"] {
            background: rgba(12, 17, 33, 0.55) !important;
            backdrop-filter: blur(16px) !important;
            border: 1px solid rgba(255, 255, 255, 0.05) !important;
            border-radius: 16px !important;
            padding: 18px 24px !important;
            box-shadow: 0 10px 30px -12px rgba(0, 0, 0, 0.5) !important;
            transition: all 0.3s cubic-bezier(0.16, 1, 0.3, 1) !important;
        }
        
        div[data-testid="stMetric"]:hover {
            transform: translateY(-3px) !important;
            border-color: rgba(79, 158, 255, 0.3) !important;
            box-shadow: 0 15px 35px -8px rgba(79, 158, 255, 0.15) !important;
        }
        
        div[data-testid="stMetricValue"] {
            color: #4f9eff !important;
            font-family: 'JetBrains Mono', monospace !important;
            font-size: 36px !important;
            font-weight: 700 !important;
            letter-spacing: -0.03em !important;
            text-shadow: 0 0 10px rgba(79, 158, 255, 0.2);
        }
        
        div[data-testid="stMetricLabel"] {
            color: #94a3b8 !important;
            font-size: 13px !important;
            font-weight: 600 !important;
            text-transform: uppercase !important;
            letter-spacing: 0.08em !important;
        }
        
        /* Premium Tabs Customization */
        button[role="tab"] {
            background-color: transparent !important;
            color: #94a3b8 !important;
            border: none !important;
            font-family: 'Syne', sans-serif !important;
            font-weight: 700 !important;
            font-size: 14px !important;
            padding: 12px 24px !important;
            transition: all 0.25s ease !important;
        }
        
        button[role="tab"][aria-selected="true"] {
            color: #4f9eff !important;
            border-bottom: 2px solid #4f9eff !important;
            text-shadow: 0 0 8px rgba(79, 158, 255, 0.3);
        }
        
        /* Neon Linear/Vercel Button Customization */
        .stButton>button {
            background: linear-gradient(135deg, #4f9eff 0%, #a78bfa 100%) !important;
            color: #060912 !important;
            border: none !important;
            border-radius: 12px !important;
            padding: 12px 28px !important;
            font-family: 'Syne', sans-serif !important;
            font-weight: 800 !important;
            letter-spacing: 0.02em !important;
            box-shadow: 0 4px 15px rgba(167, 139, 250, 0.2) !important;
            transition: all 0.3s cubic-bezier(0.16, 1, 0.3, 1) !important;
        }
        
        .stButton>button:hover {
            transform: scale(1.03) !important;
            color: #ffffff !important;
            box-shadow: 0 8px 25px rgba(167, 139, 250, 0.45) !important;
        }
        
        /* Modern Student List Hover row */
        .student-item {
            padding: 14px 20px;
            margin: 10px 0;
            background: rgba(255, 255, 255, 0.02);
            border: 1px solid rgba(255, 255, 255, 0.04);
            border-radius: 12px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            transition: all 0.25s ease;
        }
        
        .student-item:hover {
            background: rgba(255, 255, 255, 0.04);
            border-color: rgba(79, 158, 255, 0.2);
            transform: translateX(4px);
        }
        /* --- INLINE CODE TAGS (Roll No, PRN values inside HTML cards) --- */
        code {
            background: rgba(79, 158, 255, 0.12) !important;
            color: #7dd3fc !important;
            border: 1px solid rgba(79, 158, 255, 0.2) !important;
            border-radius: 6px !important;
            padding: 1px 7px !important;
            font-family: 'JetBrains Mono', monospace !important;
            font-size: 12px !important;
            font-weight: 600 !important;
        }
    </style>
""", unsafe_allow_html=True)


# ============================================================================
# B.5 DATA PERSISTENCE FUNCTIONS
# ============================================================================

DATA_FILE = "data.json"

def get_period_label(period):
    """Get readable label for a period."""
    if not period:
        return ""
    period_type = period.get("type", "lecture").capitalize()
    subject = period.get("subject", "-")
    start = period.get("start_time", "")
    end = period.get("end_time", "")
    return f"{period_type}: {subject} ({start}-{end})"


def get_period_number(period_index):
    """Convert period index to period ID (P1, P2, etc.)."""
    return f"P{period_index + 1}"


def migrate_attendance_to_new_format():
    """
    Migrate old attendance format to new period-based format.
    Old: Individual records with roll_no per entry
    New: Single record per period with attendance dict
    """
    old_records = st.session_state.period_attendance
    if not old_records:
        return
    
    # Check if already in new format (has "attendance" key with dict)
    if old_records and isinstance(old_records[0].get("attendance"), dict):
        return  # Already in new format
    
    new_records = {}  # Key: (date, day, period)
    
    for record in old_records:
        # Check if this is old format (has roll_no key)
        if "roll_no" not in record:
            # Already in new format, skip
            new_records_list = [r for r in st.session_state.period_attendance if isinstance(r.get("attendance"), dict)]
            if new_records_list:
                return
        
        date = record.get("date", datetime.now().strftime("%Y-%m-%d"))
        day = record.get("day")
        period_label = record.get("period", "")
        
        # Extract subject and type from period label
        subject = "-"
        period_type = "Lecture"
        if ":" in period_label:
            type_part, rest = period_label.split(":", 1)
            period_type = type_part.strip()
            if "(" in rest:
                subject_part, _ = rest.split("(", 1)
                subject = subject_part.strip()
        
        key = (date, day, period_label)
        
        if key not in new_records:
            # Find period index for this label
            day_periods = st.session_state.timetable.get(day, [])
            period_index = 0
            for idx, p in enumerate(day_periods):
                if get_period_label(p) == period_label:
                    period_index = idx
                    break
            
            new_records[key] = {
                "date": date,
                "day": day,
                "period": get_period_number(period_index),
                "subject": subject,
                "type": period_type.lower(),
                "attendance": {}
            }
        
        # Add student attendance
        roll_no = record.get("roll_no")
        status = record.get("status", "Absent")
        new_records[key]["attendance"][roll_no] = status
    
    # Replace old format with new format
    st.session_state.period_attendance = list(new_records.values())


def convert_bytes_to_base64(photo_bytes):
    """Convert image bytes to base64 string for JSON serialization."""
    if photo_bytes is None or not isinstance(photo_bytes, bytes):
        return None
    return base64.b64encode(photo_bytes).decode('utf-8')


def convert_base64_to_bytes(base64_string):
    """Convert base64 string back to image bytes."""
    if base64_string is None or not isinstance(base64_string, str):
        return None
    try:
        return base64.b64decode(base64_string.encode('utf-8'))
    except Exception:
        return None


def load_data():
    """Load data from JSON file and restore session state with base64 image conversion."""
    try:
        if not os.path.exists(DATA_FILE):
            return  # File doesn't exist yet, use defaults
        
        with open(DATA_FILE, 'r') as f:
            data = json.load(f)
        
        # Restore students with converted base64 photos back to bytes
        st.session_state.students = []
        for student in data.get("students", []):
            student_copy = student.copy()
            if student_copy.get("photo"):
                student_copy["photo"] = convert_base64_to_bytes(student_copy["photo"])
            st.session_state.students.append(student_copy)
        
        # Restore period attendance and migrate to new format if needed
        st.session_state.period_attendance = data.get("period_attendance", [])
        
        # Restore timetable
        st.session_state.timetable = data.get("timetable", {
            "Monday": [],
            "Tuesday": [],
            "Wednesday": [],
            "Thursday": [],
            "Friday": [],
            "Saturday": []
        })
        
        # Migrate old attendance format to new format if needed
        migrate_attendance_to_new_format()
    except Exception as e:
        st.error(f"❌ Error loading data: {str(e)}")


def save_data():
    """Save session state data to JSON file with base64-encoded images."""
    try:
        data_to_save = {
            "students": [],
            "period_attendance": st.session_state.period_attendance,
            "timetable": st.session_state.timetable
        }
        
        # Convert student photos to base64 before saving
        for student in st.session_state.students:
            student_copy = student.copy()
            if student_copy.get("photo"):
                student_copy["photo"] = convert_bytes_to_base64(student_copy["photo"])
            data_to_save["students"].append(student_copy)
        
        # Write to JSON file
        with open(DATA_FILE, 'w') as f:
            json.dump(data_to_save, f, indent=2)
    except Exception as e:
        st.error(f"❌ Error saving data: {str(e)}")


# ============================================================================
# B. SESSION STATE INITIALIZATION (MVP - Real Data Structures)
# ============================================================================

def initialize_session_state():
    """Initialize session state with empty structures for real college data."""
    
    # Load persisted data first
    load_data()
    
    # Page navigation state
    if 'page' not in st.session_state:
        st.session_state.page = "Dashboard"
    
    # Students list: List of dictionaries with college structure
    if 'students' not in st.session_state:
        st.session_state.students = []
        # Student structure: {
        #     "name": "",
        #     "roll_no": "",
        #     "prn": "",
        #     "branch": "",
        #     "class": "",
        #     "photo": ""
        # }
    
    # Attendance list: List of attendance records
    if 'attendance' not in st.session_state:
        st.session_state.attendance = []
    
    # Timetable: Dictionary by day with period types
    if 'timetable' not in st.session_state:
        st.session_state.timetable = {
            "Monday": [],
            "Tuesday": [],
            "Wednesday": [],
            "Thursday": [],
            "Friday": [],
            "Saturday": []
        }
        # Period structure: {
        #     "type": "lecture",  # or "break", "recess"
        #     "subject": "",
        #     "start_time": "",
        #     "end_time": ""
        # }
    
    # Period-based attendance: Store with day and period
    if 'period_attendance' not in st.session_state:
        st.session_state.period_attendance = []
    
    # Selected period for attendance marking
    if 'selected_day' not in st.session_state:
        st.session_state.selected_day = "Monday"
    if 'selected_period' not in st.session_state:
        st.session_state.selected_period = None
    
    # Active students for display (optional tracking)
    if 'active_students' not in st.session_state:
        st.session_state.active_students = {}
        
    if 'recognition_engine' not in st.session_state:
        st.session_state.recognition_engine = None
        
    if 'recognition_running' not in st.session_state:
        st.session_state.recognition_running = False
        
    if 'live_events' not in st.session_state:
        st.session_state.live_events = []
    
    # Ensure Saturday is added to timetable (backward compatibility)
    if "Saturday" not in st.session_state.timetable:
        st.session_state.timetable["Saturday"] = []


initialize_session_state()


# ============================================================================
# C. HELPER FUNCTIONS (MVP Logic)
# ============================================================================

def get_attendance_stats():
    """Calculate real attendance statistics from session state."""
    total_students = len(st.session_state.students)
    
    # Get today's attendance records
    today = datetime.now().strftime("%Y-%m-%d")
    today_attendance = [a for a in st.session_state.attendance if a.get("date") == today]
    
    present_count = sum(1 for a in today_attendance if a.get("status") == "Present")
    absent_count = sum(1 for a in today_attendance if a.get("status") == "Absent")
    
    attendance_rate = (present_count / total_students * 100) if total_students > 0 else 0
    
    return {
        "total": total_students,
        "present": present_count,
        "absent": absent_count,
        "late": sum(1 for a in today_attendance if a.get("status") == "Late"),
        "rate": attendance_rate
    }


def get_student_by_roll_no(roll_no):
    """Find student by roll number."""
    for student in st.session_state.students:
        if student.get("roll_no") == roll_no:
            return student
    return None


def get_student_by_prn(prn):
    """Find student by PRN (unique identifier)."""
    for student in st.session_state.students:
        if student.get("prn") == prn:
            return student
    return None


def check_overlapping_periods(day, new_start, new_end):
    """Check if new period overlaps with existing periods on the day."""
    periods = st.session_state.timetable.get(day, [])
    for period in periods:
        existing_start = period.get("start_time")
        existing_end = period.get("end_time")
        
        # Simple string comparison for time (works for HH:MM format)
        if not (new_end <= existing_start or new_start >= existing_end):
            return True  # Overlap found
    return False  # No overlap


def is_break_period(day, start_time):
    """Check if a given time slot is a break/recess on the day."""
    periods = st.session_state.timetable.get(day, [])
    for period in periods:
        if period.get("start_time") == start_time:
            period_type = period.get("type", "lecture")
            return period_type in ["break", "recess"]
    return False


def add_period_attendance(day, period, roll_no, name, status, date_str=None):
    """
    Add or update period-based attendance record.
    Stores all students for a period in a single record.
    """
    today = date_str if date_str else datetime.now().strftime("%Y-%m-%d")
    period_label = get_period_label(period)
    
    # Extract subject and type from period
    subject = period.get("subject", "-")
    period_type = period.get("type", "lecture")
    
    # Find period index to generate period ID
    day_periods = st.session_state.timetable.get(day, [])
    period_index = 0
    for idx, p in enumerate(day_periods):
        if get_period_label(p) == period_label:
            period_index = idx
            break
    
    period_id = get_period_number(period_index)
    
    # Check if record already exists for this day/period
    for record in st.session_state.period_attendance:
        if (record.get("date") == today and 
            record.get("day") == day and 
            record.get("period") == period_id):
            # Update existing record
            record["attendance"][roll_no] = status
            return True
    
    # Create new period record
    new_record = {
        "date": today,
        "day": day,
        "period": period_id,
        "subject": subject,
        "type": period_type,
        "attendance": {roll_no: status}
    }
    st.session_state.period_attendance.append(new_record)
    return True


def add_attendance_record(roll_no, name, status):
    """Add attendance record for a student."""
    today = datetime.now().strftime("%Y-%m-%d")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Check if record already exists for today
    for record in st.session_state.attendance:
        if record.get("roll_no") == roll_no and record.get("date") == today:
            record["status"] = status
            record["timestamp"] = timestamp
            return True
    
    # Add new record
    st.session_state.attendance.append({
        "roll_no": roll_no,
        "name": name,
        "status": status,
        "date": today,
        "timestamp": timestamp
    })
    return True


def get_today_attendance():
    """Get all attendance records for today."""
    today = datetime.now().strftime("%Y-%m-%d")
    return [a for a in st.session_state.attendance if a.get("date") == today]


def calculate_attendance_percentage(student_roll):
    """
    Calculate attendance percentage for a student.
    Works with new period-based attendance format.
    
    Args:
        student_roll: Roll number of the student
        
    Returns:
        float: Attendance percentage (0-100), or 0 if no records
    """
    attendance_records = []
    
    # Get all period-based attendance records for this student
    for record in st.session_state.period_attendance:
        # Check if student has attendance in this period
        if isinstance(record.get("attendance"), dict):
            if student_roll in record.get("attendance", {}):
                attendance_records.append({
                    "type": record.get("type", "lecture"),
                    "status": record.get("attendance", {}).get(student_roll, "Absent")
                })
    
    if not attendance_records:
        return 0.0
    
    # Count only lecture/lab/practical periods (ignore break/recess)
    attendance_classes = [
        a for a in attendance_records 
        if a.get("type") in ["lecture", "lab", "practical"]
    ]
    
    if not attendance_classes:
        return 0.0
    
    # Count present status (including Late and Auto-Carried states)
    present_count = sum(
        1 for a in attendance_classes 
        if a.get("status") in ["Present", "Late", "Auto-Carried"]
    )
    
    # Calculate percentage
    attendance_percentage = (present_count / len(attendance_classes)) * 100
    return round(attendance_percentage, 2)


def get_subject_wise_attendance(student_roll):
    """
    Get subject-wise attendance breakdown for a student.
    Groups attendance records by subject and calculates statistics.
    
    Args:
        student_roll: Roll number of the student
        
    Returns:
        dict: Subject-wise attendance summary with format:
            {
                "Subject1": { "total": int, "present": int, "percentage": float },
                "Subject2": { "total": int, "present": int, "percentage": float }
            }
            Returns empty dict if no records found or student_roll is None.
            
    Features:
        - Groups attendance records by subject
        - Ignores period types: break, recess
        - Counts lecture, lab, practical periods only
        - Safely handles missing data
        - Returns percentage rounded to 1 decimal place
    """
    # Handle missing or invalid input
    if not student_roll:
        return {}
    
    # Dictionary to store subject-wise stats
    subject_stats = {}
    
    # Iterate through all period attendance records
    for record in st.session_state.period_attendance:
        # Skip if not a proper record or attendance dict is missing
        if not isinstance(record, dict):
            continue
        
        # Get period type and skip break/recess periods
        period_type = record.get("type", "").lower()
        if period_type in ["break", "recess"]:
            continue
        
        # Get subject name
        subject = record.get("subject")
        if not subject:
            continue
        
        # Get attendance dictionary
        attendance_dict = record.get("attendance", {})
        if not isinstance(attendance_dict, dict):
            continue
        
        # Check if student has attendance record for this period
        if student_roll not in attendance_dict:
            continue
        
        # Initialize subject stats if not present
        if subject not in subject_stats:
            subject_stats[subject] = {
                "total": 0,
                "present": 0
            }
        
        # Increment total classes count
        subject_stats[subject]["total"] += 1
        
        # Increment present count if status is "Present", "Late", or "Auto-Carried"
        student_status = attendance_dict[student_roll]
        if student_status in ["Present", "Late", "Auto-Carried"]:
            subject_stats[subject]["present"] += 1
    
    # Calculate percentage for each subject
    result = {}
    for subject, stats in subject_stats.items():
        total = stats["total"]
        present = stats["present"]
        
        # Calculate percentage safely
        percentage = (present / total * 100) if total > 0 else 0.0
        
        result[subject] = {
            "total": total,
            "present": present,
            "percentage": round(percentage, 1)
        }
    
    return result


def get_student_attendance_summary():
    """
    Get attendance summary for all students.
    Works with new period-based attendance format.
    
    Returns:
        list: List of dictionaries with format:
            {
                "roll_no": "",
                "name": "",
                "total_classes": 0,
                "present": 0,
                "attendance_percentage": 0
            }
    """
    summary = []
    
    # Iterate through all students
    for student in st.session_state.students:
        student_roll = student.get("roll_no")
        student_name = student.get("name")
        
        # Get all attendance records for this student from period records
        total_classes = 0
        present_count = 0
        
        for record in st.session_state.period_attendance:
            # Check if this is a lecture/lab/practical period (not break/recess)
            if record.get("type") not in ["lecture", "lab", "practical"]:
                continue
            
            attendance_dict = record.get("attendance", {})
            if isinstance(attendance_dict, dict) and student_roll in attendance_dict:
                total_classes += 1
                if attendance_dict[student_roll] == "Present":
                    present_count += 1
        
        # Calculate percentage
        attendance_percentage = (
            (present_count / total_classes * 100) 
            if total_classes > 0 
            else 0
        )
        
        # Add to summary
        summary.append({
            "roll_no": student_roll,
            "name": student_name,
            "total_classes": total_classes,
            "present": present_count,
            "attendance_percentage": round(attendance_percentage, 2)
        })
    
    return summary


def export_student_report_to_excel(student_roll):
    """
    Generate and return an Excel report for a student with 3 sheets:
    - Summary: Student info and overall attendance
    - Subject-wise: Attendance breakdown by subject
    - Detailed Logs: Complete attendance history
    
    Args:
        student_roll: Roll number of the student
        
    Returns:
        bytes: Excel file as BytesIO object (in-memory), or None if student not found
        
    Features:
        - Handles missing data gracefully
        - Creates empty structured sheets if no data exists
        - Returns in-memory bytes for direct download
        - Does not save to disk
    """
    try:
        # Import pandas for Excel writing
        import pandas as pd
        from io import BytesIO
        
        # Get student info
        student = get_student_by_roll_no(student_roll)
        if not student:
            return None
        
        student_name = student.get("name", "Unknown")
        
        # Calculate overall attendance percentage
        overall_attendance = calculate_attendance_percentage(student_roll)
        
        # Get subject-wise attendance
        subject_wise_data = get_subject_wise_attendance(student_roll)
        
        # Create Excel file in memory
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # ===== SHEET 1: SUMMARY =====
            summary_data = {
                "Name": [student_name],
                "Roll No": [student_roll],
                "Overall Attendance %": [overall_attendance]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            
            # Format Summary sheet
            summary_sheet = writer.sheets["Summary"]
            for cell in summary_sheet[1]:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths for summary
            for column in summary_sheet.columns:
                max_length = 20
                column_letter = column[0].column_letter
                summary_sheet.column_dimensions[column_letter].width = max_length
            
            # ===== SHEET 2: SUBJECT-WISE =====
            if subject_wise_data:
                # Convert subject-wise dict to DataFrame
                subject_rows = []
                for subject, stats in subject_wise_data.items():
                    subject_rows.append({
                        "Subject": subject,
                        "Total Classes": stats.get("total", 0),
                        "Present": stats.get("present", 0),
                        "Attendance %": stats.get("percentage", 0.0)
                    })
                subject_df = pd.DataFrame(subject_rows)
            else:
                # Empty structure if no data
                subject_df = pd.DataFrame({
                    "Subject": [],
                    "Total Classes": [],
                    "Present": [],
                    "Attendance %": []
                })
            
            subject_df.to_excel(writer, sheet_name="Subject-wise", index=False)
            
            # Format Subject-wise sheet
            subject_sheet = writer.sheets["Subject-wise"]
            for cell in subject_sheet[1]:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths and format data
            for column in subject_sheet.columns:
                max_length = 20
                column_letter = column[0].column_letter
                subject_sheet.column_dimensions[column_letter].width = max_length
            
            # Format percentage column
            for row in subject_sheet.iter_rows(min_row=2, max_row=subject_sheet.max_row, 
                                               min_col=4, max_col=4):
                for cell in row:
                    if cell.value is not None:
                        cell.number_format = '0.0"%"'
            
            # ===== SHEET 3: DETAILED LOGS =====
            # Extract detailed attendance logs for this student
            detailed_logs = []
            for record in st.session_state.period_attendance:
                if not isinstance(record.get("attendance"), dict):
                    continue
                
                attendance_dict = record.get("attendance", {})
                if student_roll not in attendance_dict:
                    continue
                
                detailed_logs.append({
                    "Date": record.get("date", "-"),
                    "Day": record.get("day", "-"),
                    "Period": record.get("period", "-"),
                    "Subject": record.get("subject", "-"),
                    "Status": attendance_dict.get(student_roll, "-")
                })
            
            if detailed_logs:
                detailed_df = pd.DataFrame(detailed_logs)
            else:
                # Empty structure if no data
                detailed_df = pd.DataFrame({
                    "Date": [],
                    "Day": [],
                    "Period": [],
                    "Subject": [],
                    "Status": []
                })
            
            detailed_df.to_excel(writer, sheet_name="Detailed Logs", index=False)
            
            # Format Detailed Logs sheet
            logs_sheet = writer.sheets["Detailed Logs"]
            for cell in logs_sheet[1]:  # Header row
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in logs_sheet.columns:
                max_length = 18
                column_letter = column[0].column_letter
                logs_sheet.column_dimensions[column_letter].width = max_length
            
            # Center align Status column
            for row in logs_sheet.iter_rows(min_row=2, max_row=logs_sheet.max_row, 
                                            min_col=5, max_col=5):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
        
        # Return bytes
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"❌ Error generating Excel report: {str(e)}")
        return None


# ============================================================================
# C.5 MONTHLY ATTENDANCE REPORT (ERP-Level Analytics)
# ============================================================================

def get_monthly_attendance_report():
    """
    Generate comprehensive monthly attendance report for all students.
    Optimized for Excel export and ERP-level analytics dashboard.
    
    Performance: Single-pass aggregation with pre-indexed subject stats.
    Suitable for large datasets (1000+ students, 10000+ records).
    
    Returns:
        dict: {
            "rows": List[dict] - Student attendance data ready for DataFrame/Excel,
            "subject_columns": List[str] - Dynamic subject column names for UI
        }
        
    Output structure per row:
        - id: Auto-incremented index (1, 2, 3...)
        - roll_no: Student roll number (sorted)
        - name: Student name
        - branch: Student branch
        - class: Student class
        - [Subject] (Attendance %): For each subject in system
        - month_attendance_percentage: Overall average attendance
        - status: "Defaulter" (< 75%) or "OK" (≥ 75%)
        
    Features:
        - Extracts subjects dynamically (no hardcoding)
        - Skips break/recess periods
        - Missing subject data treated as 0%
        - ALL students included (even without attendance data)
        - Sorted by roll_no for stable output
        - Strict validation for dict/list safety
    """
    
    try:
        # =====================================================================
        # STEP 1: Extract all unique subjects (excluding break/recess)
        # Single pass through all records to collect subjects
        # =====================================================================
        subjects_set = set()
        
        for record in st.session_state.period_attendance:
            # Validate record structure
            if not isinstance(record, dict):
                continue
            
            # Skip break and recess periods
            period_type = record.get("type", "").lower()
            if period_type in ["break", "recess"]:
                continue
            
            # Add subject if present and valid
            subject = record.get("subject")
            if subject and subject != "-":
                subjects_set.add(subject)
        
        # Sort subjects for consistent column ordering
        subject_columns = sorted(list(subjects_set))
        
        # =====================================================================
        # STEP 2: Pre-aggregate attendance data by student and subject
        # Optimized structure: {roll_no: {subject: {"total": int, "present": int}}}
        # Avoids O(n*m) lookups later when building rows
        # =====================================================================
        student_subject_stats = {}
        
        for record in st.session_state.period_attendance:
            # Validate record structure
            if not isinstance(record, dict):
                continue
            
            # Skip break/recess periods
            period_type = record.get("type", "").lower()
            if period_type in ["break", "recess"]:
                continue
            
            subject = record.get("subject")
            if not subject or subject == "-":
                continue
            
            # Get attendance dictionary with validation
            attendance_dict = record.get("attendance")
            if not isinstance(attendance_dict, dict):
                continue
            
            # Process each student in this period
            for roll_no, status in attendance_dict.items():
                # Initialize student if not exists
                if roll_no not in student_subject_stats:
                    student_subject_stats[roll_no] = {}
                
                # Initialize subject if not exists
                if subject not in student_subject_stats[roll_no]:
                    student_subject_stats[roll_no][subject] = {"total": 0, "present": 0}
                
                # Increment counters (only count Present status)
                student_subject_stats[roll_no][subject]["total"] += 1
                if status == "Present":
                    student_subject_stats[roll_no][subject]["present"] += 1
        
        # =====================================================================
        # STEP 3: Build report rows for all students with O(n) complexity
        # =====================================================================
        rows = []
        
        for student in st.session_state.students:
            roll_no = student.get("roll_no")
            name = student.get("name", "-")
            branch = student.get("branch", "-")
            class_name = student.get("class", "-")
            
            # Get subject-wise attendance for this student (from pre-aggregated data)
            student_subjects = student_subject_stats.get(roll_no, {})
            
            # Build row with static columns
            row = {
                "id": None,  # Will be set after sorting
                "roll_no": roll_no,
                "name": name,
                "branch": branch,
                "class": class_name
            }
            
            # Add subject-wise attendance columns dynamically
            total_attendance_list = []
            for subject in subject_columns:
                subject_col_name = f"{subject} (Attendance %)"
                
                # Get subject stats or default to 0% if missing
                subject_stats = student_subjects.get(subject, {"total": 0, "present": 0})
                total = subject_stats.get("total", 0)
                present = subject_stats.get("present", 0)
                
                # Calculate percentage safely
                if total > 0:
                    percentage = round((present / total) * 100, 1)
                else:
                    percentage = 0.0
                
                row[subject_col_name] = f"{percentage}%"
                total_attendance_list.append(percentage)
            
            # Calculate overall monthly attendance percentage
            if total_attendance_list:
                month_attendance_percentage = round(
                    sum(total_attendance_list) / len(total_attendance_list),
                    1
                )
            else:
                month_attendance_percentage = 0.0
            
            row["month_attendance_percentage"] = f"{month_attendance_percentage}%"
            
            # Determine status based on numeric value
            if month_attendance_percentage >= 75:
                row["status"] = "OK"
            else:
                row["status"] = "Defaulter"
            
            rows.append(row)
        
        # =====================================================================
        # STEP 4: Sort by roll_no for stable, consistent output
        # =====================================================================
        rows.sort(key=lambda x: str(x.get("roll_no", "")))
        
        # Re-index IDs after sorting
        for idx, row in enumerate(rows, 1):
            row["id"] = idx
        
        # =====================================================================
        # STEP 5: Build formatted subject columns list for UI
        # =====================================================================
        formatted_subject_columns = [
            f"{subject} (Attendance %)" 
            for subject in subject_columns
        ]
        
        return {
            "rows": rows,
            "subject_columns": formatted_subject_columns
        }
        
    except Exception as e:
        st.error(f"❌ Error generating monthly attendance report: {str(e)}")
        return {
            "rows": [],
            "subject_columns": []
        }


# ============================================================================
# C.6 MONTHLY STUDENT MATRIX REPORT (Dashboard Ready)
# ============================================================================

def get_monthly_student_matrix_report(month=None):
    """
    Generate monthly student matrix report with dynamic subject columns.
    Optimized for Streamlit dashboard display as interactive matrix.
    
    Args:
        month: Optional month filter (YYYY-MM format). If None, uses current month.
        
    Returns:
        dict: {
            "data": DataFrame with student x subject matrix,
            "summary": Overall statistics,
            "subject_list": List of subjects,
            "defaulters": List of defaulter students
        }
        
    Matrix Structure:
        - Rows: Students (sorted by roll_no)
        - Columns: ID, Roll No, Name, Branch, Class, [Subject]..., Avg %, Status
        - Cell values: Attendance % or status
        
    Features:
        - Auto-extracts subjects dynamically
        - Color-coded status indicators
        - Summary statistics (total students, defaulters, avg attendance)
        - Ready for Excel export and dashboard display
    """
    
    try:
        # Get base report data
        report_data = get_monthly_attendance_report()
        rows = report_data.get("rows", [])
        subject_columns = report_data.get("subject_columns", [])
        
        if not rows:
            return {
                "data": pd.DataFrame(),
                "summary": {
                    "total_students": 0,
                    "defaulters_count": 0,
                    "average_attendance": 0.0
                },
                "subject_list": [],
                "defaulters": []
            }
        
        # Convert to DataFrame for easier manipulation
        df = pd.DataFrame(rows)
        
        # Calculate summary statistics
        total_students = len(df)
        defaulters = df[df["status"] == "Defaulter"]
        defaulters_count = len(defaulters)
        
        # Extract numeric average attendance percentage (remove % sign)
        df["numeric_attendance"] = df["month_attendance_percentage"].str.rstrip('%').astype(float)
        average_attendance = df["numeric_attendance"].mean()
        
        summary = {
            "total_students": total_students,
            "defaulters_count": defaulters_count,
            "average_attendance": round(average_attendance, 1),
            "ok_count": total_students - defaulters_count
        }
        
        # Build defaulters list with detail
        defaulters_list = []
        for _, row in defaulters.iterrows():
            defaulters_list.append({
                "roll_no": row["roll_no"],
                "name": row["name"],
                "attendance_percentage": row["month_attendance_percentage"],
                "branch": row["branch"],
                "class": row["class"]
            })
        
        return {
            "data": df,
            "summary": summary,
            "subject_list": subject_columns,
            "defaulters": defaulters_list
        }
        
    except Exception as e:
        st.error(f"❌ Error generating matrix report: {str(e)}")
        return {
            "data": pd.DataFrame(),
            "summary": {
                "total_students": 0,
                "defaulters_count": 0,
                "average_attendance": 0.0
            },
            "subject_list": [],
            "defaulters": []
        }


# ============================================================================
# D. DASHBOARD PAGE
# ============================================================================

def show_dashboard():
    """Display the main dashboard with college attendance metrics."""
    
    # Inject Custom CSS with Animations and Modern Dark-Theme styling
    st.markdown("""
        <style>
            @keyframes fadeIn {
                from { opacity: 0; transform: translateY(20px); }
                to { opacity: 1; transform: translateY(0); }
            }
            @keyframes pulse {
                0% { transform: scale(1); box-shadow: 0 0 0 0 rgba(16, 185, 129, 0.4); }
                70% { transform: scale(1.02); box-shadow: 0 0 0 10px rgba(16, 185, 129, 0); }
                100% { transform: scale(1); box-shadow: 0 0 0 0 rgba(16, 185, 129, 0); }
            }
            .animated-header {
                animation: fadeIn 0.8s ease-out forwards;
                background: linear-gradient(135deg, #1e293b 0%, #0f172a 100%);
                border-radius: 16px;
                padding: 24px;
                margin-bottom: 24px;
                border: 1px solid rgba(255, 255, 255, 0.08);
                box-shadow: 0 10px 30px rgba(0,0,0,0.3);
            }
            .animated-card {
                animation: fadeIn 0.6s ease-out forwards;
                background: linear-gradient(135deg, rgba(30, 41, 59, 0.7) 0%, rgba(15, 23, 42, 0.8) 100%);
                border: 1px solid rgba(255, 255, 255, 0.08);
                border-radius: 16px;
                padding: 20px;
                margin-bottom: 15px;
                box-shadow: 0 8px 24px rgba(0,0,0,0.2);
                transition: transform 0.3s ease, box-shadow 0.3s ease, border-color 0.3s ease;
                backdrop-filter: blur(12px);
            }
            .animated-card:hover {
                transform: translateY(-5px);
                box-shadow: 0 12px 30px rgba(0,0,0,0.3);
                border-color: rgba(16, 185, 129, 0.3);
            }
            .active-period-card {
                animation: pulse 2s infinite;
                border: 2px solid #10b981 !important;
                background: linear-gradient(135deg, rgba(16, 185, 129, 0.12) 0%, rgba(15, 23, 42, 0.85) 100%) !important;
            }
            .gradient-text {
                background: linear-gradient(90deg, #10b981, #3b82f6);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
                font-weight: 800;
            }
        </style>
    """, unsafe_allow_html=True)
    
    # 1. Header Section
    now = datetime.now()
    today_str = now.strftime("%A, %B %d, %Y")
    today_name = now.strftime("%A")
    current_time_str = now.strftime("%I:%M %p")
    
    st.markdown(f"""
        <div class="animated-header">
            <div style="display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 15px;">
                <div>
                    <h1 style="margin: 0; font-size: 32px;"><span class="gradient-text">📊 AI Attendance Dashboard</span></h1>
                    <p style="margin: 5px 0 0 0; color: rgba(255,255,255,0.6); font-size: 16px;">Smart College Attendance & Timetable Analytics</p>
                </div>
                <div style="text-align: right; background: rgba(255,255,255,0.05); padding: 10px 20px; border-radius: 12px; border: 1px solid rgba(255,255,255,0.1);">
                    <span style="font-size: 14px; color: #10b981; font-weight: 600; display: block;">🕒 LIVE SYSTEM TIME</span>
                    <span style="font-size: 20px; font-weight: 700; color: #fff;">{today_str} • {current_time_str}</span>
                </div>
            </div>
        </div>
    """, unsafe_allow_html=True)
    
    stats = get_attendance_stats()
    
    # 2. Metric Overview Row
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""
            <div class="animated-card" style="text-align: center; background: linear-gradient(135deg, rgba(30, 41, 59, 0.8) 0%, rgba(15, 23, 42, 0.9) 100%);">
                <span style="font-size: 13px; color: rgba(255,255,255,0.5); font-weight:600; display:block; margin-bottom:5px;">TOTAL STUDENTS</span>
                <span style="font-size: 36px; font-weight: 800; color: #fff;">{stats['total']}</span>
            </div>
        """, unsafe_allow_html=True)
    with col2:
        st.markdown(f"""
            <div class="animated-card" style="text-align: center; background: linear-gradient(135deg, rgba(16, 185, 129, 0.1) 0%, rgba(15, 23, 42, 0.9) 100%); border-color: rgba(16, 185, 129, 0.2);">
                <span style="font-size: 13px; color: #10b981; font-weight:600; display:block; margin-bottom:5px;">✅ PRESENT TODAY</span>
                <span style="font-size: 36px; font-weight: 800; color: #10b981;">{stats['present']}</span>
            </div>
        """, unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
            <div class="animated-card" style="text-align: center; background: linear-gradient(135deg, rgba(239, 68, 68, 0.1) 0%, rgba(15, 23, 42, 0.9) 100%); border-color: rgba(239, 68, 68, 0.2);">
                <span style="font-size: 13px; color: #ef4444; font-weight:600; display:block; margin-bottom:5px;">❌ ABSENT TODAY</span>
                <span style="font-size: 36px; font-weight: 800; color: #ef4444;">{stats['absent']}</span>
            </div>
        """, unsafe_allow_html=True)
    with col4:
        st.markdown(f"""
            <div class="animated-card" style="text-align: center; background: linear-gradient(135deg, rgba(59, 130, 246, 0.1) 0%, rgba(15, 23, 42, 0.9) 100%); border-color: rgba(59, 130, 246, 0.2);">
                <span style="font-size: 13px; color: #3b82f6; font-weight:600; display:block; margin-bottom:5px;">📈 ATTENDANCE RATE</span>
                <span style="font-size: 36px; font-weight: 800; color: #3b82f6;">{stats['rate']:.1f}%</span>
            </div>
        """, unsafe_allow_html=True)
        
    st.markdown("---")
    
    # 3. Dynamic Current Day Timetable Timeline
    st.markdown("### 📅 Timetable & Schedule Search")
    
    # Check if today's classes are finished to automatically roll over to next day's timetable
    now_dt = datetime.now()
    today_str_check = now_dt.strftime("%Y-%m-%d")
    day_str_check = now_dt.strftime("%A")
    if today_str_check in st.session_state.timetable:
        today_periods_check = st.session_state.timetable[today_str_check]
    else:
        today_periods_check = st.session_state.timetable.get(day_str_check, [])
        
    default_date = now_dt
    was_rolled_over = False
    if today_periods_check:
        try:
            latest_end_time = datetime.strptime("00:00", "%H:%M").time()
            for p in today_periods_check:
                p_end = datetime.strptime(p.get("end_time", "00:00"), "%H:%M").time()
                if p_end > latest_end_time:
                    latest_end_time = p_end
            if now_dt.time() > latest_end_time:
                default_date = now_dt + timedelta(days=1)
                was_rolled_over = True
        except:
            pass
            
    search_date = st.date_input(
        "🔍 Search Timetable by Date",
        default_date,
        key="dash_timetable_search_date"
    )
    search_date_str = search_date.strftime("%Y-%m-%d")
    search_day_name = search_date.strftime("%A")
    
    if was_rolled_over and search_date.strftime("%Y-%m-%d") == (now_dt + timedelta(days=1)).strftime("%Y-%m-%d"):
        st.success(f"🌅 Today's classes are finished! Automatically showing **tomorrow's timetable** to keep you ahead.")
    
    if search_date_str in st.session_state.timetable:
        today_periods = st.session_state.timetable[search_date_str]
        st.info(f"✨ **Custom Date-Specific Schedule** active for **{search_date_str}** ({search_day_name})!")
    else:
        today_periods = st.session_state.timetable.get(search_day_name, [])
        st.info(f"📅 Showing standard timetable schedule for **{search_day_name}**, **{search_date_str}**")
    
    if not today_periods:
        st.info(f"🏖️ No classes scheduled for {search_day_name}, {search_date_str}. Enjoy your day off!")
    else:
        # Display periods in a beautiful custom timeline
        for idx, p in enumerate(today_periods):
            subj = p.get('subject', 'Unknown')
            per_id = p.get('period', 'Unknown')
            p_type = p.get('type', 'lecture').upper()
            start_time = p.get('start_time', '00:00')
            end_time = p.get('end_time', '00:00')
            room = p.get('room', 'N/A')
            
            # Determine if this specific class is currently active based on system time (only if selected date is today)
            is_active = False
            if search_date_str == datetime.now().strftime("%Y-%m-%d"):
                try:
                    start = datetime.strptime(start_time, "%H:%M").time()
                    end = datetime.strptime(end_time, "%H:%M").time()
                    is_active = (start <= now.time() <= end)
                except Exception:
                    pass
            
            card_class = "animated-card active-period-card" if is_active else "animated-card"
            active_badge = "<span style='background:#10b981; color:#fff; padding:3px 10px; border-radius:12px; font-size:11px; font-weight:bold; margin-left:12px; letter-spacing:0.5px;'>ACTIVE NOW</span>" if is_active else ""
            
            st.markdown(f"""
                <div class="{card_class}" style="padding: 18px 24px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 15px;">
                    <div style="display: flex; align-items: center; gap: 20px;">
                        <div style="background: rgba(59, 130, 246, 0.1); padding: 12px 18px; border-radius: 12px; text-align: center; border: 1px solid rgba(59, 130, 246, 0.2); min-width: 80px;">
                            <span style="font-size: 20px; font-weight: 800; color: #3b82f6; display: block;">{per_id}</span>
                        </div>
                        <div>
                            <h4 style="margin: 0; font-size: 18px; font-weight: 700; color: #fff;">{subj} {active_badge}</h4>
                            <span style="font-size: 13px; color: rgba(255,255,255,0.5); font-weight: 500;">{p_type} • Room/Lab: <code>{room}</code></span>
                        </div>
                    </div>
                    <div style="text-align: right;">
                        <span style="font-size: 16px; font-weight: 700; color: #10b981; display: block;">🕒 {start_time} - {end_time}</span>
                    </div>
                </div>
            """, unsafe_allow_html=True)
            
    st.markdown("---")
    
    # 4. Searchable & Filterable Student Directory with Full Metadata
    st.markdown("### 👥 Registered Students Directory")
    
    if not st.session_state.students:
        st.warning("⚠️ No students registered in the college database yet. Go to 'Manage Students' to register students.")
    else:
        # Create Search & Filter Controls
        filter_col1, filter_col2 = st.columns([3, 1])
        with filter_col1:
            search_query = st.text_input("🔍 Search Students", placeholder="Search by name, roll number, or PRN...", key="search_query")
        with filter_col2:
            unique_branches = sorted(list(set(student.get("branch", "Other") for student in st.session_state.students)))
            branch_filter = st.selectbox("Branch Filter", ["All Branches"] + unique_branches)
            
        # Apply Filters
        filtered_students = st.session_state.students
        if search_query:
            search_query = search_query.lower()
            filtered_students = [s for s in filtered_students if (
                search_query in s.get("name", "").lower() or
                search_query in s.get("roll_no", "").lower() or
                search_query in s.get("prn", "").lower()
            )]
        if branch_filter != "All Branches":
            filtered_students = [s for s in filtered_students if s.get("branch") == branch_filter]
            
        # Display Student Profiles with Hover Animations
        if not filtered_students:
            st.info("No registered students match your search filters.")
        else:
            for student in filtered_students:
                roll_no = student.get("roll_no")
                name = student.get("name")
                prn = student.get("prn", "N/A")
                branch = student.get("branch", "N/A")
                student_class = student.get("class", "N/A")
                photo_bytes = student.get("photo")
                
                # Calculate live cumulative attendance percentage
                attendance_pct = calculate_attendance_percentage(roll_no)
                
                # Color code attendance based on 75% college rule
                if attendance_pct >= 75:
                    status_badge = f"<span style='background:#10b981; color:#fff; padding:6px 12px; border-radius:12px; font-size:14px; font-weight:800; letter-spacing:0.5px;'>{attendance_pct}%</span>"
                elif attendance_pct >= 50:
                    status_badge = f"<span style='background:#f59e0b; color:#fff; padding:6px 12px; border-radius:12px; font-size:14px; font-weight:800; letter-spacing:0.5px;'>{attendance_pct}%</span>"
                else:
                    status_badge = f"<span style='background:#ef4444; color:#fff; padding:6px 12px; border-radius:12px; font-size:14px; font-weight:800; letter-spacing:0.5px;'>{attendance_pct}%</span>"
                
                st.markdown(f"""
                    <div class="animated-card" style="margin-bottom: 12px;">
                        <div style="display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 15px;">
                            <div style="display: flex; align-items: center; gap: 15px;">
                                <div style="font-size: 32px; background: rgba(255,255,255,0.05); padding: 8px; border-radius: 50%; min-width: 48px; text-align: center;">👤</div>
                                <div>
                                    <h4 style="margin: 0; font-size: 18px; font-weight: 700; color: #fff;">{name}</h4>
                                    <span style="font-size: 13px; color: rgba(255,255,255,0.5); font-weight: 500;">Roll No: <span style="background:rgba(79,158,255,0.15); color:#7dd3fc; padding:1px 7px; border-radius:6px; font-family:'JetBrains Mono',monospace; font-size:12px; font-weight:700; border:1px solid rgba(79,158,255,0.25);">{roll_no}</span> &nbsp;•&nbsp; PRN: <span style="background:rgba(79,158,255,0.15); color:#7dd3fc; padding:1px 7px; border-radius:6px; font-family:'JetBrains Mono',monospace; font-size:12px; font-weight:700; border:1px solid rgba(79,158,255,0.25);">{prn}</span></span>
                                </div>
                            </div>
                            <div style="display: flex; gap: 40px; align-items: center;">
                                <div>
                                    <span style="font-size: 12px; color: rgba(255,255,255,0.4); font-weight: 600; display: block; text-align: right;">CLASS & DEPT</span>
                                    <span style="font-size: 14px; font-weight: 700; color: #3b82f6; display: block; text-align: right;">{student_class} - {branch}</span>
                                </div>
                                <div style="text-align: right; min-width: 100px;">
                                    <span style="font-size: 12px; color: rgba(255,255,255,0.4); font-weight: 600; display: block; margin-bottom: 4px; text-align: right;">OVERALL</span>
                                    {status_badge}
                                </div>
                            </div>
                        </div>
                    </div>
                """, unsafe_allow_html=True)



# ============================================================================
# E. ADD STUDENT PAGE
# ============================================================================

def show_add_student():
    """Add Student page - Collect student details and add to session state."""
    st.title("➕ Add Student")
    
    st.write("Add new students to the college system. Each student must have unique roll number and PRN.")
    st.markdown("---")
    
    # ===== STUDENT DETAILS SECTION =====
    st.subheader("👤 Student Details")
    col1, col2 = st.columns(2)
    
    with col1:
        student_name = st.text_input(
            "Student Name",
            placeholder="Enter full name",
            key="add_student_name"
        )
    
    with col2:
        roll_number = st.text_input(
            "Roll Number",
            placeholder="e.g., 101",
            key="add_student_roll"
        )
    
    col3, col4 = st.columns(2)
    
    with col3:
        prn = st.text_input(
            "PRN (Permanent Registration Number)",
            placeholder="e.g., 2024001",
            key="add_student_prn"
        )
    
    with col4:
        st.text("")  # Spacer
        st.text("")
    
    # ===== ACADEMIC DETAILS SECTION =====
    st.subheader("📚 Academic Details")
    col5, col6 = st.columns(2)
    
    with col5:
        branch = st.selectbox(
            "Branch",
            ["Computer Science", "Electronics", "Mechanical", "Civil", "Electrical", "Other"],
            key="add_student_branch"
        )
    
    with col6:
        class_name = st.text_input(
            "Class/Semester",
            placeholder="e.g., FE (First Year), SE, TE, BE",
            key="add_student_class"
        )
    
    # ===== PHOTO UPLOAD SECTION =====
    st.subheader("📷 Photo Upload")
    uploaded_file = st.file_uploader(
        "Upload Student Photo (Optional)",
        type=["jpg", "jpeg", "png"],
        key="add_student_photo"
    )
    
    photo_bytes = None
    if uploaded_file is not None:
        photo_bytes = uploaded_file.read()
        if isinstance(photo_bytes, bytes):
            st.image(photo_bytes, width=150, caption="Uploaded Photo")
        else:
            st.warning("⚠️ Old image format not supported. Please re-upload.")
    
    # Validation and Add button
    st.markdown("---")
    if st.button("➕ Add Student", key="btn_add_student"):
        # Validation
        if not student_name or not roll_number or not prn or not branch or not class_name:
            st.error("❌ Please fill in all required fields (Name, Roll, PRN, Branch, Class)")
        elif get_student_by_roll_no(roll_number):
            st.error(f"❌ Roll number '{roll_number}' already exists!")
        elif get_student_by_prn(prn):
            st.error(f"❌ PRN '{prn}' already exists!")
        else:
            # Add student to session state
            new_student = {
                "name": student_name,
                "roll_no": roll_number,
                "prn": prn,
                "branch": branch,
                "class": class_name,
                "photo": photo_bytes
            }
            st.session_state.students.append(new_student)
            save_data()  # Persist changes
            st.success(f"✅ Student '{student_name}' (Roll: {roll_number}, PRN: {prn}) added successfully!")

    
    st.markdown("---")
    st.subheader(f"📋 All Students ({len(st.session_state.students)})")
    
    if len(st.session_state.students) == 0:
        st.info("No students added yet. Add a student using the form above.")
    else:
        # Display students in a table-like format
        for i, student in enumerate(st.session_state.students, 1):
            col1, col2, col3 = st.columns([2, 3, 1])
            
            with col1:
                st.write(f"**{student['name']}**")
                st.caption(f"Roll: {student['roll_no']} | PRN: {student['prn']}")
            
            with col2:
                st.write(f"**{student['branch']}** | {student['class']}")
            
            with col3:
                if student['photo']:
                    st.caption("📷 Photo")




# ============================================================================
# E.5 STUDENT DIRECTORY PAGE (NEW)
# ============================================================================

def show_manage_students():
    """Display all students and allow adding/removing students with AI embedding generation."""
    st.title("👥 Manage Students")
    st.write("View, add, and remove students from the AI Attendance System.")
    st.markdown("---")
    
    tab1, tab2 = st.tabs(["📋 All Students Directory", "➕ Add New Student"])
    
    with tab1:
        if len(st.session_state.students) == 0:
            st.markdown("""
                <div style="background:rgba(6,9,18,0.8); border:2px dashed rgba(79,158,255,0.2); border-radius:16px; padding:40px; text-align:center; margin-top:20px;">
                    <div style="font-size:40px; margin-bottom:12px;">👥</div>
                    <div style="color:#94a3b8; font-size:15px;">No students registered yet. Add your first student in the <strong style="color:#4f9eff;">Add New Student</strong> tab.</div>
                </div>
            """, unsafe_allow_html=True)
        else:
            # ── KPI Strip ──
            total_s = len(st.session_state.students)
            unique_branches = len(set(s.get('branch', 'Unknown') for s in st.session_state.students))
            students_with_photo = sum(1 for s in st.session_state.students if s.get('photo'))
            
            k1, k2, k3 = st.columns(3)
            with k1: st.metric("👥 Total Students", total_s)
            with k2: st.metric("🏢 Branches", unique_branches)
            with k3: st.metric("📷 With Photos", students_with_photo)
            
            st.markdown("---")

            # ── Student Cards Grid ──
            st.markdown('<p style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#64748b;margin-bottom:16px;">📋 Student Directory</p>', unsafe_allow_html=True)
            
            for idx, student in enumerate(st.session_state.students):
                roll_no = student.get('roll_no', 'N/A')
                name = student.get('name', 'Unknown')
                prn = student.get('prn', 'N/A')
                branch = student.get('branch', 'N/A')
                cls = student.get('class', 'N/A')
                att_pct = calculate_attendance_percentage(roll_no)

                if att_pct >= 75:
                    att_color = "#34d399"
                    att_bg = "rgba(52,211,153,0.1)"
                elif att_pct >= 50:
                    att_color = "#fb923c"
                    att_bg = "rgba(251,146,60,0.1)"
                else:
                    att_color = "#f472b6"
                    att_bg = "rgba(244,114,182,0.1)"

                has_photo = "✅" if student.get('photo') else "❌"

                st.markdown(f"""
                    <div class="animated-card" style="padding:16px 20px; margin-bottom:10px; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:12px;">
                        <div style="display:flex; align-items:center; gap:16px; flex:1; min-width:200px;">
                            <div style="width:40px; height:40px; border-radius:50%; background:linear-gradient(135deg,#4f9eff,#a78bfa); display:flex; align-items:center; justify-content:center; font-weight:800; color:#fff; font-size:16px; flex-shrink:0;">{idx}</div>
                            <div>
                                <div style="font-weight:700; font-size:15px; color:#ffffff;">{name}</div>
                                <div style="font-size:12px; color:#64748b; margin-top:2px;">
                                    <span style="background:rgba(79,158,255,0.12); color:#7dd3fc; padding:1px 6px; border-radius:5px; font-family:monospace; font-size:11px; font-weight:700;">Roll {roll_no}</span>
                                    &nbsp;•&nbsp;
                                    <span style="background:rgba(79,158,255,0.12); color:#7dd3fc; padding:1px 6px; border-radius:5px; font-family:monospace; font-size:11px; font-weight:700;">PRN {prn}</span>
                                </div>
                            </div>
                        </div>
                        <div style="display:flex; gap:24px; align-items:center; flex-wrap:wrap;">
                            <div style="text-align:center;">
                                <div style="font-size:10px; color:#64748b; text-transform:uppercase; letter-spacing:0.06em;">Branch</div>
                                <div style="font-size:13px; font-weight:700; color:#a78bfa;">{branch}</div>
                            </div>
                            <div style="text-align:center;">
                                <div style="font-size:10px; color:#64748b; text-transform:uppercase; letter-spacing:0.06em;">Class</div>
                                <div style="font-size:13px; font-weight:700; color:#94a3b8;">{cls}</div>
                            </div>
                            <div style="text-align:center;">
                                <div style="font-size:10px; color:#64748b; text-transform:uppercase; letter-spacing:0.06em;">Photo</div>
                                <div style="font-size:13px;">{has_photo}</div>
                            </div>
                            <div style="text-align:center; background:{att_bg}; padding:8px 14px; border-radius:12px; border:1px solid {att_color}33;">
                                <div style="font-size:10px; color:#64748b; text-transform:uppercase; letter-spacing:0.06em;">Attendance</div>
                                <div style="font-size:16px; font-weight:800; color:{att_color};">{att_pct}%</div>
                            </div>
                        </div>
                    </div>
                """, unsafe_allow_html=True)

            st.markdown("---")

            # ── Student Detail Panel ──
            st.markdown('<p style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#64748b;margin-bottom:12px;">🔍 Student Detail & Management</p>', unsafe_allow_html=True)

            sel_col, _ = st.columns([0.25, 0.75])
            with sel_col:
                selected_index = st.number_input("Select Student ID:", min_value=0, max_value=len(st.session_state.students) - 1, value=0, step=1, key="dir_student_index")

            selected_student = st.session_state.students[int(selected_index)]
            attendance_pct = calculate_attendance_percentage(selected_student.get('roll_no'))

            summary_data = {"total": 0, "present": 0}
            for record in get_student_attendance_summary():
                if record['roll_no'] == selected_student['roll_no']:
                    summary_data = {"total": record['total_classes'], "present": record['present']}
                    break

            if attendance_pct >= 75:
                badge_color = "#34d399"; badge_bg = "rgba(52,211,153,0.12)"
                status_label = "✅ Good Standing"
            elif attendance_pct >= 50:
                badge_color = "#fb923c"; badge_bg = "rgba(251,146,60,0.12)"
                status_label = "⚠️ Warning Zone"
            else:
                badge_color = "#f472b6"; badge_bg = "rgba(244,114,182,0.12)"
                status_label = "🚨 Defaulter Risk"

            d1, d2, d3 = st.columns([0.28, 0.38, 0.34])

            with d1:
                st.markdown('<div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:8px;">📷 Photo</div>', unsafe_allow_html=True)
                if selected_student.get('photo') and isinstance(selected_student['photo'], bytes):
                    st.image(selected_student['photo'], use_container_width=True, caption=selected_student['name'])
                else:
                    st.markdown("""
                        <div style="background:rgba(255,255,255,0.03);border:1px dashed rgba(255,255,255,0.1);border-radius:12px;padding:30px;text-align:center;color:#475569;">
                            <div style="font-size:36px;">👤</div>
                            <div style="font-size:12px;margin-top:6px;">No Photo</div>
                        </div>
                    """, unsafe_allow_html=True)

            with d2:
                st.markdown('<div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:8px;">📋 Student Info</div>', unsafe_allow_html=True)
                st.markdown(f"""
                    <div class="animated-card" style="padding:20px;">
                        <div style="margin-bottom:12px;">
                            <div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:0.06em;">Full Name</div>
                            <div style="font-size:16px;font-weight:800;color:#ffffff;margin-top:2px;">{selected_student['name']}</div>
                        </div>
                        <div style="display:grid; grid-template-columns:1fr 1fr; gap:12px;">
                            <div>
                                <div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:0.06em;">Roll No</div>
                                <div style="font-size:14px;font-weight:700;color:#7dd3fc;font-family:monospace;margin-top:2px;">{selected_student['roll_no']}</div>
                            </div>
                            <div>
                                <div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:0.06em;">PRN</div>
                                <div style="font-size:14px;font-weight:700;color:#7dd3fc;font-family:monospace;margin-top:2px;">{selected_student['prn']}</div>
                            </div>
                            <div>
                                <div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:0.06em;">Branch</div>
                                <div style="font-size:14px;font-weight:700;color:#a78bfa;margin-top:2px;">{selected_student['branch']}</div>
                            </div>
                            <div>
                                <div style="font-size:10px;color:#64748b;text-transform:uppercase;letter-spacing:0.06em;">Class</div>
                                <div style="font-size:14px;font-weight:700;color:#94a3b8;margin-top:2px;">{selected_student['class']}</div>
                            </div>
                        </div>
                    </div>
                """, unsafe_allow_html=True)

            with d3:
                st.markdown('<div style="font-size:12px;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.06em;margin-bottom:8px;">📊 Attendance</div>', unsafe_allow_html=True)
                st.markdown(f"""
                    <div class="animated-card" style="padding:20px; text-align:center;">
                        <div style="font-size:42px; font-weight:900; color:{badge_color}; font-family:'JetBrains Mono',monospace;">{attendance_pct}%</div>
                        <div style="background:{badge_bg}; color:{badge_color}; border-radius:20px; padding:4px 12px; display:inline-block; font-size:11px; font-weight:700; margin:8px 0 16px 0;">{status_label}</div>
                        <div style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:8px; margin-top:4px;">
                            <div style="background:rgba(255,255,255,0.03); border-radius:10px; padding:8px;">
                                <div style="font-size:10px; color:#64748b;">Total</div>
                                <div style="font-size:18px; font-weight:800; color:#ffffff;">{summary_data['total']}</div>
                            </div>
                            <div style="background:rgba(52,211,153,0.08); border-radius:10px; padding:8px;">
                                <div style="font-size:10px; color:#64748b;">Present</div>
                                <div style="font-size:18px; font-weight:800; color:#34d399;">{summary_data['present']}</div>
                            </div>
                            <div style="background:rgba(244,114,182,0.08); border-radius:10px; padding:8px;">
                                <div style="font-size:10px; color:#64748b;">Absent</div>
                                <div style="font-size:18px; font-weight:800; color:#f472b6;">{summary_data['total'] - summary_data['present']}</div>
                            </div>
                        </div>
                    </div>
                """, unsafe_allow_html=True)

            st.markdown("")
            _, _, btn_col = st.columns(3)
            with btn_col:
                if st.button("🗑️ Remove Student", use_container_width=True, key="dir_remove_btn_merged"):
                    removed_roll = selected_student['roll_no']
                    removed_name = selected_student['name']
                    st.session_state.students.pop(int(selected_index))
                    st.session_state.attendance = [a for a in st.session_state.attendance if a.get("roll_no") != removed_roll]
                    save_data()
                    st.success(f"✅ Student '{removed_name}' removed successfully!")
                    st.rerun()

    with tab2:
        st.markdown("### Add New Student & Capture Face")
        with st.form("add_student_form_merged"):
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                name = st.text_input("Full Name*", placeholder="e.g. John Doe")
            with col2:
                roll_no = st.text_input("Roll Number*", placeholder="e.g. 101")
            with col3:
                prn = st.text_input("College PRN*", placeholder="e.g. PRN12345")
            with col4:
                branch = st.selectbox("Branch", ["Computer Science", "Electronics", "Mechanical", "Civil", "Electrical", "Other"])
                
            student_class = st.selectbox("Class", ["FY", "SY", "TY", "B.Tech", "FY B.Tech", "SY B.Tech", "TY B.Tech"])
                
            st.markdown("### 📷 Face Image Input")
            st.info("Please provide a face image by EITHER uploading a file OR taking a camera photo.")
            
            uploaded_file = st.file_uploader("1. Upload Student Face Image", type=["jpg", "jpeg", "png"])
            st.markdown("**-- OR --**")
            camera_photo = st.camera_input("2. Take Student Photo")
            
            submit = st.form_submit_button("Save Student & Generate Embeddings", type="primary", use_container_width=True)
            
            if submit:
                if not name or not roll_no or not prn:
                    st.error("Name, Roll Number, and PRN are required fields.")
                elif get_student_by_roll_no(roll_no):
                    st.error(f"❌ Roll number '{roll_no}' already exists!")
                elif get_student_by_prn(prn):
                    st.error(f"❌ PRN '{prn}' already exists!")
                elif camera_photo is None and uploaded_file is None:
                    st.error("Please provide a face image (upload or capture).")
                else:
                    try:
                        import cv2
                        import numpy as np
                        from PIL import Image
                        import os, json
                        from datetime import datetime
                        
                        img_source = camera_photo if camera_photo is not None else uploaded_file
                        photo_bytes = img_source.getvalue()
                        
                        image = Image.open(img_source)
                        img_array = np.array(image)
                        
                        if len(img_array.shape) == 2:
                            img_bgr = cv2.cvtColor(img_array, cv2.COLOR_GRAY2BGR)
                        elif img_array.shape[2] == 4:
                            img_bgr = cv2.cvtColor(img_array, cv2.COLOR_RGBA2BGR)
                        else:
                            img_bgr = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
                        
                        # Face duplicate check
                        if 'recognition_engine' not in st.session_state or st.session_state.recognition_engine is None:
                            from vision.face_recognition_engine import RealTimeFaceRecognitionEngine
                            st.session_state.recognition_engine = RealTimeFaceRecognitionEngine()
                        
                        engine = st.session_state.recognition_engine
                        faces = engine.face_analyzer.get(img_bgr)
                        
                        if len(faces) == 0:
                            st.error("❌ No face detected in the image. Please try again with a clearer photo.")
                        elif len(faces) > 1:
                            st.error("❌ Multiple faces detected. Please ensure only one student is in the frame.")
                        else:
                            embedding = faces[0].embedding
                            match = engine.match_face(embedding)
                            
                            # Using 0.45 as a strict threshold for "same person"
                            if match and match['similarity'] > 0.45:
                                st.error(f"❌ This face is already registered in the system under: {match['name']} (Roll: {match['roll_no']}) with {match['similarity']*100:.1f}% confidence!")
                            else:
                                resized_img = cv2.resize(img_bgr, (1000, 1000), interpolation=cv2.INTER_LINEAR)
                                
                                safe_name = "".join([c for c in name if c.isalnum() or c==' ']).rstrip()
                                dir_name = f"{roll_no}_{safe_name.replace(' ', '_')}"
                                save_dir = os.path.join("dataset", dir_name)
                                os.makedirs(save_dir, exist_ok=True)
                                
                                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                                filename = f"face_{roll_no}_{timestamp}.jpg"
                                filepath = os.path.join(save_dir, filename)
                                
                                cv2.imwrite(filepath, resized_img, [cv2.IMWRITE_JPEG_QUALITY, 100])
                                
                                new_student = {
                                    "name": name,
                                    "roll_no": roll_no,
                                    "prn": prn,
                                    "branch": branch,
                                    "class": student_class,
                                    "photo": photo_bytes
                                }
                                st.session_state.students.append(new_student)
                                save_data()
                                
                                meta_path = os.path.join(save_dir, "metadata.json")
                                with open(meta_path, 'w') as f:
                                    json.dump({"name": name, "roll_no": roll_no, "prn": prn, "branch": branch, "class": student_class, "added_on": timestamp}, f)
                                
                                st.success(f"✅ Successfully added student '{name}' and saved image to dataset!")
                                st.info("🔄 Generating AI Face Embeddings... This may take a moment.")
                                
                                from vision.generate_embeddings import generate_face_embeddings
                                result = generate_face_embeddings(verbose=False)
                                
                                if result.get('status') in ['success', 'partial_success']:
                                    st.success("✅ AI Embeddings updated successfully! The student will now be recognized.")
                                    st.session_state.recognition_engine = None
                                else:
                                    st.error(f"⚠️ Warning: Failed to generate embeddings: {result.get('error', 'Unknown Error')}")
                            
                    except Exception as e:
                        st.error(f"Failed to process image: {e}")

def show_student_directory():
    """Display all students in a hybrid table + detail panel layout."""
    st.title("📖 Student Directory")
    
    st.write("View and manage all students in the system with their complete information and attendance analytics.")
    st.markdown("---")
    
    # ===== ADD STUDENT BUTTON =====
    col1, col2 = st.columns([0.15, 0.85])
    with col1:
        if st.button("➕ Add Student", use_container_width=True):
            st.session_state.page = "Add Student"
            st.rerun()
    
    st.markdown("---")
    
    # ===== EMPTY STATE =====
    if len(st.session_state.students) == 0:
        st.info("📌 No students in the system yet. Click the 'Add Student' button above to get started.")
        return
    
    # ===== STATS SECTION =====
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("👥 Total Students", len(st.session_state.students))
    with col2:
        unique_branches = len(set(s.get('branch', 'Unknown') for s in st.session_state.students))
        st.metric("🏢 Branches", unique_branches)
    with col3:
        students_with_photo = sum(1 for s in st.session_state.students if s.get('photo'))
        st.metric("📷 With Photos", students_with_photo)
    
    st.markdown("---")
    
    # ===== PRIMARY TABLE VIEW =====
    st.subheader("📊 Student Directory Table")
    st.write("Click on a row number below to view detailed information:")
    
    # Create table data with all required columns
    table_data = []
    for idx, student in enumerate(st.session_state.students):
        attendance_pct = calculate_attendance_percentage(student.get('roll_no'))
        table_data.append({
            "Index": idx,
            "Roll No": student.get('roll_no'),
            "Name": student.get('name'),
            "PRN": student.get('prn'),
            "Branch": student.get('branch'),
            "Class": student.get('class'),
            "Photo": "✅" if student.get('photo') else "❌",
            "Attendance %": f"{attendance_pct}%"
        })
    
    # Display as professional dataframe
    st.dataframe(
        table_data,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Index": st.column_config.NumberColumn("ID", width="small"),
            "Roll No": st.column_config.TextColumn("Roll No", width="small"),
            "Name": st.column_config.TextColumn("Name", width="medium"),
            "PRN": st.column_config.TextColumn("PRN", width="medium"),
            "Branch": st.column_config.TextColumn("Branch", width="medium"),
            "Class": st.column_config.TextColumn("Class", width="small"),
            "Photo": st.column_config.TextColumn("Photo", width="small"),
            "Attendance %": st.column_config.TextColumn("Attendance %", width="small")
        }
    )
    
    st.markdown("---")
    
    # ===== ROW SELECTION & DETAIL PANEL =====
    st.subheader("🔍 Student Details & Management")
    
    col1, col2 = st.columns([0.3, 0.7])
    
    with col1:
        selected_index = st.number_input(
            "Select Student (by ID):",
            min_value=0,
            max_value=len(st.session_state.students) - 1,
            value=0,
            step=1,
            key="dir_student_index"
        )
    
    # Get selected student
    selected_student = st.session_state.students[int(selected_index)]
    attendance_pct = calculate_attendance_percentage(selected_student.get('roll_no'))
    
    # Display detail panel
    detail_col1, detail_col2, detail_col3 = st.columns([0.35, 0.35, 0.3])
    
    with detail_col1:
        st.write("### � Student Photo")
        if selected_student.get('photo'):
            if isinstance(selected_student['photo'], bytes):
                st.image(selected_student['photo'], width=200, caption=f"{selected_student['name']}'s Photo")
            else:
                st.warning("⚠️ Old image format not supported. Please re-upload.")
        else:
            st.info("📌 No photo uploaded")
    
    with detail_col2:
        st.write("### 📋 Student Information")
        st.write(f"**Name:** {selected_student['name']}")
        st.write(f"**Roll No:** {selected_student['roll_no']}")
        st.write(f"**PRN:** {selected_student['prn']}")
        st.write(f"**Branch:** {selected_student['branch']}")
        st.write(f"**Class:** {selected_student['class']}")
    
    with detail_col3:
        st.write("### 📊 Attendance")
        st.metric("Attendance %", f"{attendance_pct}%")
        
        # Get total and present counts
        summary = get_student_attendance_summary()
        for record in summary:
            if record['roll_no'] == selected_student['roll_no']:
                st.write(f"**Total:** {record['total_classes']}")
                st.write(f"**Present:** {record['present']}")
                st.write(f"**Absent:** {record['total_classes'] - record['present']}")
                break
    
    st.markdown("---")
    
    # ===== ACTION BUTTONS =====
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("🔄 View Full Attendance", use_container_width=True):
            st.info(f"📌 Attendance records for {selected_student['name']} (Roll: {selected_student['roll_no']})")
            # Display period-based attendance for this student
            student_attendance = []
            for record in st.session_state.period_attendance:
                attendance_dict = record.get("attendance", {})
                if isinstance(attendance_dict, dict) and selected_student['roll_no'] in attendance_dict:
                    status = attendance_dict[selected_student['roll_no']]
                    student_attendance.append({
                        "date": record.get("date"),
                        "day": record.get("day"),
                        "period": record.get("period"),
                        "subject": record.get("subject"),
                        "type": record.get("type"),
                        "status": status
                    })
            
            if student_attendance:
                for record in student_attendance:
                    st.write(f"**{record['date']}** | {record['day']} | {record['period']} ({record['subject']}) | {record['type'].upper()} | Status: **{record['status']}**")
            else:
                st.write("No attendance records found.")
    
    with col2:
        st.write("")  # Spacer
    
    with col3:
        if st.button("🗑️ Remove Student", use_container_width=True, key="dir_remove_btn"):
            # Remove student from list
            removed_name = st.session_state.students[int(selected_index)]['name']
            st.session_state.students.pop(int(selected_index))
            save_data()  # Persist changes
            st.success(f"✅ Student '{removed_name}' removed successfully!")
            st.rerun()



# ============================================================================
# F. REMOVE STUDENT PAGE



# ============================================================================
# F. REMOVE STUDENT PAGE
# ============================================================================

def show_remove_student():
    """Remove Student page - Remove students with college details."""
    st.title("➖ Remove Student")
    
    st.write("Remove students from the system. All related attendance records will also be removed.")
    st.markdown("---")
    
    if len(st.session_state.students) == 0:
        st.info("No students to remove. Add students first using the 'Add Student' page.")
        return
    
    # Display current students
    st.subheader(f"Current Students ({len(st.session_state.students)})")
    
    # Create roll number options
    roll_numbers = [s["roll_no"] for s in st.session_state.students]
    
    selected_roll = st.selectbox(
        "Select student to remove:",
        roll_numbers,
        format_func=lambda x: f"{x} - {get_student_by_roll_no(x)['name']}"
    )
    
    # Show selected student details
    if selected_roll:
        student = get_student_by_roll_no(selected_roll)
        st.info(f"""
        **Student to Remove:**
        - Name: {student['name']}
        - Roll: {student['roll_no']}
        - PRN: {student['prn']}
        - Branch: {student['branch']}
        - Class: {student['class']}
        """)
    
    if st.button("🗑️ Remove Selected Student"):
        # Find and remove student
        for i, student in enumerate(st.session_state.students):
            if student["roll_no"] == selected_roll:
                removed_student = st.session_state.students.pop(i)
                st.session_state.attendance = [
                    a for a in st.session_state.attendance 
                    if a.get("roll_no") != selected_roll
                ]
                save_data()  # Persist changes
                st.success(f"✅ Student '{removed_student['name']}' (PRN: {removed_student['prn']}) removed successfully!")
                st.info("✓ All related attendance records also removed.")
                break
    
    st.markdown("---")
    st.subheader("📋 Remaining Students")
    if st.session_state.students:
        for i, student in enumerate(st.session_state.students, 1):
            st.write(f"{i}. **{student['name']}** (Roll: {student['roll_no']}, PRN: {student['prn']}) - {student['branch']}")


# ============================================================================
# G. TIMETABLE PAGE
# ============================================================================

def show_timetable():
    """Timetable page - Display and manage class schedule with period types."""
    st.title("📅 Timetable Management")
    
    st.write("Manage your college class schedule. Add lectures, breaks, and recess periods.")
    st.markdown("---")
    
    # ===== RESET OPTIONS =====
    st.subheader("🔄 Reset Timetable Options")
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("🗑️ Reset Day Timetable", key="btn_reset_day"):
            st.warning("⚠️ Select the day to reset in the form below, then click this button again.")
    
    with col2:
        if st.button("🗑️ Reset Full Timetable", key="btn_reset_all"):
            if st.session_state.get("confirm_reset_all"):
                st.session_state.timetable = {
                    "Monday": [],
                    "Tuesday": [],
                    "Wednesday": [],
                    "Thursday": [],
                    "Friday": [],
                    "Saturday": []
                }
                save_data()  # Persist changes
                st.success("✅ Full timetable reset successfully!")
                st.session_state.confirm_reset_all = False
                st.rerun()
            else:
                st.error("⚠️ Click the 'Confirm Reset' button below to reset the full timetable.")
    
    col1, col2 = st.columns(2)
    with col1:
        reset_day = st.selectbox("Day to Reset", list(st.session_state.timetable.keys()), key="reset_day_select")
        if st.button("✅ Reset Selected Day", key="btn_confirm_reset_day"):
            st.session_state.timetable[reset_day] = []
            save_data()  # Persist changes
            st.success(f"✅ {reset_day} timetable reset successfully!")
            st.rerun()
    
    with col2:
        st.session_state.confirm_reset_all = st.checkbox("I confirm to reset the entire timetable", key="confirm_reset_check")
    
    st.markdown("---")
    
    # ===== ADD PERIOD SECTION =====
    st.subheader("➕ Add New Period")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        tt_type_select = st.radio("Schedule For", ["Weekly Day", "Specific Date"], horizontal=True, key="tt_sched_type")
        if tt_type_select == "Weekly Day":
            day = st.selectbox("Select Day", ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"], key="tt_day")
        else:
            cust_date = st.date_input("Select Custom Date", datetime.now(), key="tt_cust_date")
            day = cust_date.strftime("%Y-%m-%d")
    
    with col2:
        period_type = st.selectbox(
            "Period Type",
            ["Lecture", "Break", "Recess", "Lab", "Practical"],
            key="tt_type"
        )
    
    with col3:
        period_name = st.text_input("Period Name*", placeholder="e.g. P1, P2", key="tt_period_name")
        
    with col4:
        subject = st.text_input(
            "Subject",
            placeholder="e.g. Mathematics",
            key="tt_subject"
        )
    
    col5, col6, _ = st.columns([1, 1, 2])
    
    with col5:
        start_time = st.text_input("Start Time", placeholder="09:00", key="tt_start")
    
    with col6:
        end_time = st.text_input("End Time", placeholder="10:00", key="tt_end")
    
    # Add Period Button
    if st.button("➕ Add Period", key="btn_add_period", type="primary"):
        if not start_time or not end_time or not period_name:
            st.error("❌ Please fill in start time, end time, and Period Name")
        elif period_type in ["Lecture", "Lab", "Practical"] and not subject:
            st.error("❌ Subject is required for lecture, lab, and practical periods")
        elif check_overlapping_periods(day, start_time, end_time):
            st.error(f"❌ Period overlaps with existing period on {day}")
        else:
            new_period = {
                "period": period_name.strip(),
                "type": period_type.lower(),
                "subject": subject.strip() if period_type not in ["Break", "Recess"] else period_type,
                "start_time": start_time.strip(),
                "end_time": end_time.strip()
            }
            if day not in st.session_state.timetable:
                st.session_state.timetable[day] = []
            st.session_state.timetable[day].append(new_period)
            save_data()  # Persist changes
            st.success(f"✅ {period_type} period added for {day} ({start_time}-{end_time})!")
            st.rerun()
    
    st.markdown("---")
    
    # ===== WEEKLY & CUSTOM SCHEDULES SECTION =====
    st.subheader("📊 Class Schedules")
    
    weekly_days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
    custom_dates = sorted([k for k in st.session_state.timetable.keys() if k not in weekly_days])
    
    st.markdown("### 🗓️ Weekly Days Schedule")
    cols = st.columns(3)
    for idx, day in enumerate(weekly_days):
        with cols[idx % 3]:
            st.markdown(f"#### 📅 {day}")
            periods = st.session_state.timetable.get(day, [])
            if not periods:
                st.info("Free Day 🏖️")
            else:
                for p_idx, p in enumerate(periods):
                    p_name = p.get('period', f"#{p_idx+1}")
                    p_type = p.get('type', 'lecture').capitalize()
                    subj = p.get('subject', '-')
                    start = p.get('start_time', '')
                    end = p.get('end_time', '')
                    
                    accent_color = "#4f9eff" if p_type in ["Lecture", "Lab", "Practical"] else "#fb923c"
                    p_type_label = p_type.upper()
                    
                    st.markdown(f"""
                        <div class="animated-card" style="padding: 16px; margin-bottom: 8px; border-left: 4px solid {accent_color} !important;">
                            <h4 style="margin: 0 0 4px 0; font-size: 15px; font-weight: 800; color: #ffffff;">{p_name} - {subj}</h4>
                            <span style="font-size: 12px; color: #94a3b8; font-weight: 600; display: block; margin-bottom: 8px;">🕒 {start} to {end}</span>
                            <span style="background: rgba(255,255,255,0.06); color: {accent_color}; padding: 3px 8px; border-radius: 6px; font-size: 10px; font-weight: 800; text-transform: uppercase; letter-spacing: 0.05em;">{p_type_label}</span>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    if st.button("🗑️ Delete Period", key=f"del_{day}_{p_idx}", use_container_width=True):
                        st.session_state.timetable[day].pop(p_idx)
                        save_data()
                        st.rerun()
                                
    if custom_dates:
        st.markdown("---")
        st.markdown("### 🌟 Date-Specific Custom Schedules")
        cols_custom = st.columns(3)
        for idx, day in enumerate(custom_dates):
            with cols_custom[idx % 3]:
                st.markdown(f"#### 📅 {day}")
                periods = st.session_state.timetable.get(day, [])
                if not periods:
                    st.info("No customized periods")
                else:
                    for p_idx, p in enumerate(periods):
                        p_name = p.get('period', f"#{p_idx+1}")
                        p_type = p.get('type', 'lecture').capitalize()
                        subj = p.get('subject', '-')
                        start = p.get('start_time', '')
                        end = p.get('end_time', '')
                        
                        accent_color = "#4f9eff" if p_type in ["Lecture", "Lab", "Practical"] else "#fb923c"
                        p_type_label = p_type.upper()
                        
                        st.markdown(f"""
                            <div class="animated-card" style="padding: 16px; margin-bottom: 8px; border-left: 4px solid {accent_color} !important;">
                                <h4 style="margin: 0 0 4px 0; font-size: 15px; font-weight: 800; color: #ffffff;">{p_name} - {subj}</h4>
                                <span style="font-size: 12px; color: #94a3b8; font-weight: 600; display: block; margin-bottom: 8px;">🕒 {start} to {end}</span>
                                <span style="background: rgba(255,255,255,0.06); color: {accent_color}; padding: 3px 8px; border-radius: 6px; font-size: 10px; font-weight: 800; text-transform: uppercase; letter-spacing: 0.05em;">{p_type_label}</span>
                            </div>
                        """, unsafe_allow_html=True)
                        
                        if st.button("🗑️ Delete Period", key=f"del_{day}_{p_idx}", use_container_width=True):
                            st.session_state.timetable[day].pop(p_idx)
                            if not st.session_state.timetable[day]:
                                del st.session_state.timetable[day]
                            save_data()
                            st.rerun()
    
    st.markdown("---")
    
    # ===== PERIOD TYPE LEGEND =====
    st.subheader("📌 Period Type Legend")
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.success("**Lecture**: Regular class (attendance required)")
    
    with col2:
        st.warning("**Break**: Short break (no attendance)")
    
    with col3:
        st.warning("**Recess**: Long recess (no attendance)")
    
    with col4:
        st.info("**Lab**: Practical lab session (attendance required)")
    
    with col5:
        st.info("**Practical**: Practical session (attendance required)")


# ============================================================================
# H. LIVE ATTENDANCE PAGE
# ============================================================================

def show_live_attendance():
    """Live Attendance page - Mark attendance linked to specific periods."""
    st.title("🎥 Live Attendance Management (Period-Based)")
    
    st.write("Mark attendance for students linked to specific lecture/lab/practical periods.")
    st.markdown("---")
    
    if len(st.session_state.students) == 0:
        st.warning("⚠️ No students in the system. Add students first!")
        return
    
    # ===== SELECT DATE & PERIOD SECTION =====
    st.subheader("📅 Select Date & Period")
    
    col_d1, col_d2 = st.columns(2)
    
    with col_d1:
        selected_date = st.date_input(
            "Select Attendance Date",
            datetime.now(),
            key="att_date"
        )
        selected_date_str = selected_date.strftime("%Y-%m-%d")
        selected_day = selected_date.strftime("%A")
        st.info(f"📅 Detected Weekday: **{selected_day}**")
    
    with col_d2:
        if selected_date_str in st.session_state.timetable:
            day_periods = st.session_state.timetable[selected_date_str]
            st.success(f"✨ Custom schedule loaded for date **{selected_date_str}**!")
            selected_day = selected_date_str
        else:
            day_periods = st.session_state.timetable.get(selected_day, [])
        # Filter out break and recess periods
        markable_periods = [p for p in day_periods if p.get("type") not in ["break", "recess"]]
        period_labels = [get_period_label(p) for p in markable_periods]
        
        if period_labels:
            selected_period_label = st.selectbox(
                "Select Period (Lectures/Labs/Practicals Only)",
                period_labels,
                key="att_period"
            )
            # Find the period object
            selected_period = None
            for p in markable_periods:
                if get_period_label(p) == selected_period_label:
                    selected_period = p
                    break
        else:
            st.error(f"❌ No lecture/lab/practical periods scheduled for {selected_day}!")
            selected_period = None
    
    if not selected_period:
        st.info("⏳ Please select a calendar date and a scheduled period to mark attendance.")
        return
    
    # Display selected period details
    st.success(f"📌 Selected Class: {selected_period_label} on {selected_date_str} ({selected_day})")
    
    st.markdown("---")
    
    # ===== MARK ATTENDANCE SECTION =====
    st.subheader("✅ Mark Attendance")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        selected_roll = st.selectbox(
            "Select Student",
            [s["roll_no"] for s in st.session_state.students],
            format_func=lambda x: f"{x} - {get_student_by_roll_no(x)['name']}",
            key="att_student"
        )
    
    with col2:
        status = st.selectbox("Status", ["Present", "Absent", "Late"], key="att_status")
    
    with col3:
        if st.button("✅ Mark Attendance", key="btn_mark_att"):
            student = get_student_by_roll_no(selected_roll)
            add_period_attendance(selected_day, selected_period, selected_roll, student["name"], status, date_str=selected_date_str)
            save_data()  # Persist changes
            st.success(f"✅ Attendance marked for {selected_date_str}: {student['name']} ({student['roll_no']}) - {status}")
            st.rerun()
    
    st.markdown("---")
    
    # ===== PERIOD ATTENDANCE STATISTICS =====
    st.subheader(f"📊 {selected_period_label} ({selected_date_str}) - Attendance Summary")
    
    today = selected_date_str
    period_index = 0
    day_periods = st.session_state.timetable.get(selected_day, [])
    for idx, p in enumerate(day_periods):
        if get_period_label(p) == selected_period_label:
            period_index = idx
            break
    period_id = get_period_number(period_index)
    
    # Find the period record for selected date
    period_record = None
    for record in st.session_state.period_attendance:
        if (record.get("date") == today and 
            record.get("day") == selected_day and 
            record.get("period") == period_id):
            period_record = record
            break
    
    # Get attendance data from the period record
    attendance_dict = period_record.get("attendance", {}) if period_record else {}
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Students", len(st.session_state.students))
    with col2:
        present_count = sum(1 for status in attendance_dict.values() if status == "Present")
        st.metric("✅ Present", present_count)
    with col3:
        absent_count = sum(1 for status in attendance_dict.values() if status == "Absent")
        st.metric("❌ Absent", absent_count)
    with col4:
        late_count = sum(1 for status in attendance_dict.values() if status == "Late")
        st.metric("⏱️ Late", late_count)
    
    st.markdown("---")
    
    # ===== DETAILED ATTENDANCE LIST =====
    st.subheader("📋 Attendance Details for This Period")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**✅ Present:**")
        present_list = [(roll, status) for roll, status in attendance_dict.items() if status == "Present"]
        if present_list:
            for roll_no, status in present_list:
                student = get_student_by_roll_no(roll_no)
                student_name = student.get("name", roll_no) if student else roll_no
                st.write(f"✅ {student_name} ({roll_no})")
        else:
            st.info("No present records yet")
    
    with col2:
        st.write("**❌ Absent / ⏱️ Late:**")
        absent_list = [(roll, status) for roll, status in attendance_dict.items() if status in ["Absent", "Late"]]
        if absent_list:
            for roll_no, status in absent_list:
                student = get_student_by_roll_no(roll_no)
                student_name = student.get("name", roll_no) if student else roll_no
                status_emoji = "❌" if status == "Absent" else "⏱️"
                st.write(f"{status_emoji} {student_name} ({roll_no}) - {status}")
        else:
            st.info("No absent/late records yet")
    
    st.markdown("---")
    
    # ===== UNMARKED STUDENTS =====
    st.subheader("⏳ Students Not Yet Marked for This Period")
    
    marked_rolls = set(attendance_dict.keys())
    unmarked_students = [s for s in st.session_state.students if s["roll_no"] not in marked_rolls]
    
    if unmarked_students:
        for s in unmarked_students:
            col1, col2 = st.columns([3, 1])
            with col1:
                st.write(f"⏳ {s['name']} ({s['roll_no']}) - {s['branch']}")
            with col2:
                st.caption(f"{s['class']}")
    else:
        st.success("✅ All students marked for this period!")


def show_reports_page():
    """Display real-time daily attendance matrix and export advanced excel reports."""

    st.markdown("""
        <div style="margin-bottom:28px;">
            <h2 style="margin:0; font-size:30px; font-weight:900; background:linear-gradient(135deg,#a78bfa,#4f9eff); -webkit-background-clip:text; -webkit-text-fill-color:transparent;">
                📊 Attendance Reports
            </h2>
            <p style="color:#64748b; font-size:13px; margin:4px 0 0 0;">Real-time live matrix, date-wise, week-wise & month-wise attendance with Excel export</p>
        </div>
    """, unsafe_allow_html=True)

    if len(st.session_state.students) == 0:
        st.markdown("""
            <div style="background:rgba(6,9,18,0.8); border:2px dashed rgba(251,146,60,0.3); border-radius:16px; padding:40px; text-align:center;">
                <div style="font-size:40px; margin-bottom:12px;">⚠️</div>
                <div style="color:#fb923c; font-size:15px; font-weight:700;">No students in the system.</div>
                <div style="color:#64748b; font-size:13px; margin-top:6px;">Add students first before generating reports.</div>
            </div>
        """, unsafe_allow_html=True)
        return

    # Reload fresh data
    try:
        with open("data.json", "r") as _f:
            _live = json.load(_f)
        _live_tt = _live.get("timetable", {})
        st.session_state.period_attendance = _live.get("period_attendance", st.session_state.period_attendance)
    except Exception:
        _live_tt = st.session_state.timetable

    today_str = datetime.now().strftime("%Y-%m-%d")
    day_str = datetime.now().strftime("%A")

    tab1, tab2 = st.tabs(["📋 Today's Live Matrix", "📥 Export Reports"])

    # ─────────────────────────────────────────────
    # TAB 1: TODAY'S LIVE MATRIX
    # ─────────────────────────────────────────────
    with tab1:
        timetable = _live_tt.get(today_str) or _live_tt.get(day_str, [])

        st.markdown(f"""
            <div style="display:flex; align-items:center; justify-content:space-between; margin-bottom:20px; flex-wrap:wrap; gap:12px;">
                <div>
                    <div style="font-size:18px; font-weight:800; color:#ffffff;">Live Attendance — {day_str}</div>
                    <div style="font-size:12px; color:#64748b;">{today_str}</div>
                </div>
                <div style="background:rgba(52,211,153,0.1); border:1px solid rgba(52,211,153,0.3); border-radius:20px; padding:6px 14px;">
                    <span style="color:#34d399; font-size:12px; font-weight:700;">● LIVE</span>
                </div>
            </div>
        """, unsafe_allow_html=True)

        if not timetable:
            st.markdown(f"""
                <div style="background:rgba(6,9,18,0.8); border:2px dashed rgba(79,158,255,0.2); border-radius:16px; padding:50px; text-align:center;">
                    <div style="font-size:40px; margin-bottom:12px;">🏖️</div>
                    <div style="color:#94a3b8; font-size:15px; font-weight:700;">No Periods Scheduled Today</div>
                    <div style="color:#64748b; font-size:13px; margin-top:6px;">{day_str} ({today_str}) — Add periods in the Timetable section.</div>
                </div>
            """, unsafe_allow_html=True)
        else:
            try:
                with open("data/erp_attendance.json", "r") as f:
                    erp_data = json.load(f)
            except Exception:
                erp_data = {}
            today_erp = erp_data.get(today_str, {})

            # Period summary strip
            period_cols = st.columns(min(len(timetable), 6))
            for i, p in enumerate(timetable):
                subj = p.get('subject', '?')
                per_id = p.get('period', f'P{i+1}')
                start = p.get('start_time', '')
                end = p.get('end_time', '')
                per_records = today_erp.get(subj, {}).get(per_id, {})
                present_count = sum(1 for v in per_records.values() if v.get("status") in ["Present", "Late", "Auto-Carried"])
                total = len(st.session_state.students)
                pct = round(present_count / total * 100) if total > 0 else 0
                accent = "#34d399" if pct >= 75 else "#fb923c" if pct >= 50 else "#f472b6"
                with period_cols[i % len(period_cols)]:
                    st.markdown(f"""
                        <div class="animated-card" style="padding:14px; text-align:center; border-left:4px solid {accent} !important; margin-bottom:8px;">
                            <div style="font-size:10px; color:#64748b; text-transform:uppercase; letter-spacing:0.08em;">{per_id}</div>
                            <div style="font-size:14px; font-weight:800; color:#ffffff; margin:4px 0;">{subj[:12]}</div>
                            <div style="font-size:11px; color:#64748b;">{start}–{end}</div>
                            <div style="font-size:22px; font-weight:900; color:{accent}; margin-top:6px;">{present_count}/{total}</div>
                            <div style="font-size:10px; color:{accent};">{pct}% Present</div>
                        </div>
                    """, unsafe_allow_html=True)

            st.markdown("---")

            # Build full student×period matrix
            report_data = []
            for idx, student in enumerate(st.session_state.students, 1):
                roll_no = student.get("roll_no", "")
                row = {"#": idx, "Roll No": roll_no, "Name": student.get("name", ""), "Branch": student.get("branch", ""), "Class": student.get("class", "")}
                for p in timetable:
                    subj = p.get('subject', 'Unknown')
                    per_id = p.get('period', 'Unknown')
                    col_name = f"{subj} [{per_id}]"
                    rec = today_erp.get(subj, {}).get(per_id, {}).get(roll_no)
                    if rec and rec.get("status") in ["Present", "Late", "Auto-Carried"]:
                        row[col_name] = "✅"
                        row[f"_time_{col_name}"] = rec.get("entry_time", "")
                    else:
                        row[col_name] = "❌"
                        row[f"_time_{col_name}"] = "-"
                report_data.append(row)

            df = pd.DataFrame(report_data)
            display_cols = [c for c in df.columns if not c.startswith("_time_")]

            # ── Build HTML table ──
            period_cols_hdr = [c for c in display_cols if c not in ["#","Roll No","Name","Branch","Class"]]

            header_cells = "".join([
                f'<th style="padding:10px 14px;text-align:center;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.08em;color:#64748b;white-space:nowrap;border-bottom:2px solid rgba(79,158,255,0.15);">{c}</th>'
                for c in ["#","Roll No","Name","Branch","Class"]
            ])
            for c in period_cols_hdr:
                header_cells += f'<th style="padding:10px 14px;text-align:center;font-size:10px;font-weight:700;color:#4f9eff;white-space:nowrap;border-bottom:2px solid rgba(79,158,255,0.25);">{c}</th>'

            body_rows = ""
            for _, row in df[display_cols].iterrows():
                bg = "rgba(255,255,255,0.01)" if int(row["#"]) % 2 == 0 else "transparent"
                cells = "".join([
                    f'<td style="padding:10px 14px;color:#94a3b8;font-size:12px;white-space:nowrap;">{row[c]}</td>'
                    for c in ["#","Roll No","Name","Branch","Class"]
                ])
                for c in period_cols_hdr:
                    v = row[c]
                    if v == "✅":
                        badge = '<span style="background:rgba(52,211,153,0.15);color:#34d399;border:1px solid rgba(52,211,153,0.3);border-radius:6px;padding:3px 10px;font-size:11px;font-weight:700;">✅ Present</span>'
                    elif v == "⏱️":
                        badge = '<span style="background:rgba(251,146,60,0.15);color:#fb923c;border:1px solid rgba(251,146,60,0.3);border-radius:6px;padding:3px 10px;font-size:11px;font-weight:700;">⏱ Late</span>'
                    else:
                        badge = '<span style="background:rgba(244,114,182,0.1);color:#f472b6;border:1px solid rgba(244,114,182,0.2);border-radius:6px;padding:3px 10px;font-size:11px;font-weight:700;">❌ Absent</span>'
                    cells += f'<td style="padding:8px 14px;text-align:center;">{badge}</td>'
                body_rows += f'<tr style="background:{bg};border-bottom:1px solid rgba(255,255,255,0.03);">{cells}</tr>'

            matrix_html = f"""
            <div style="background:rgba(12,17,33,0.8);border:1px solid rgba(79,158,255,0.12);border-radius:14px;overflow:hidden;margin-bottom:8px;">
                <div style="padding:14px 18px 10px;border-bottom:1px solid rgba(79,158,255,0.1);display:flex;align-items:center;gap:8px;">
                    <span style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#64748b;">📋 Full Attendance Matrix</span>
                    <span style="background:rgba(79,158,255,0.1);color:#4f9eff;border-radius:20px;padding:2px 10px;font-size:10px;font-weight:700;">{len(df)} Students</span>
                    <span style="background:rgba(52,211,153,0.1);color:#34d399;border-radius:20px;padding:2px 10px;font-size:10px;font-weight:700;">{len(period_cols_hdr)} Periods</span>
                </div>
                <div style="overflow-x:auto;">
                    <table style="width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;">
                        <thead><tr style="background:rgba(6,9,18,0.6);">{header_cells}</tr></thead>
                        <tbody>{body_rows}</tbody>
                    </table>
                </div>
            </div>
            """
            st.markdown(matrix_html, unsafe_allow_html=True)

            st.markdown("---")

            # Per-student period detail
            st.markdown('<p style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#64748b;margin-bottom:10px;">🕒 Student Period Drill-Down</p>', unsafe_allow_html=True)
            roll_list = df["Roll No"].tolist()
            if roll_list:
                sel_roll = st.selectbox("Select Student", roll_list,
                    format_func=lambda x: f"{x} — {df[df['Roll No']==x]['Name'].values[0]}",
                    key="reports_ts_sel")
                if sel_roll:
                    sdata = df[df["Roll No"] == sel_roll].iloc[0]
                    drill_cols = st.columns(max(len(timetable), 1))
                    for i, p in enumerate(timetable):
                        subj = p.get('subject', 'Unknown')
                        per_id = p.get('period', 'Unknown')
                        col_name = f"{subj} [{per_id}]"
                        val = sdata.get(col_name, "❌")
                        t_val = sdata.get(f"_time_{col_name}", "-")
                        is_present = val == "✅"
                        accent = "#34d399" if is_present else "#f472b6"
                        with drill_cols[i % len(drill_cols)]:
                            st.markdown(f"""
                                <div class="animated-card" style="padding:14px; text-align:center; border-left:4px solid {accent} !important;">
                                    <div style="font-size:11px; color:#64748b;">{per_id}</div>
                                    <div style="font-size:13px; font-weight:800; color:#ffffff; margin:4px 0;">{subj[:14]}</div>
                                    <div style="font-size:28px; margin:8px 0;">{val}</div>
                                    <div style="font-size:11px; color:#64748b;">{"⏱ " + t_val if is_present else "Absent"}</div>
                                </div>
                            """, unsafe_allow_html=True)

    # ─────────────────────────────────────────────
    # TAB 2: EXPORT REPORTS — 5 scopes
    # ─────────────────────────────────────────────
    with tab2:
        st.markdown("""
            <div style="background:linear-gradient(135deg,rgba(79,158,255,0.06),rgba(167,139,250,0.06));border:1px solid rgba(79,158,255,0.15);border-radius:16px;padding:20px;margin-bottom:20px;">
                <div style="font-size:15px;font-weight:800;color:#ffffff;margin-bottom:12px;">📥 Export Attendance Reports</div>
                <div style="display:grid;grid-template-columns:repeat(5,1fr);gap:8px;">
                    <div style="background:rgba(79,158,255,0.08);border:1px solid rgba(79,158,255,0.2);border-radius:10px;padding:10px;text-align:center;">
                        <div style="font-size:16px;">📅</div><div style="font-size:11px;font-weight:700;color:#4f9eff;margin-top:4px;">Day-wise</div>
                        <div style="font-size:10px;color:#64748b;margin-top:2px;">All Mondays / Fridays etc.</div>
                    </div>
                    <div style="background:rgba(167,139,250,0.08);border:1px solid rgba(167,139,250,0.2);border-radius:10px;padding:10px;text-align:center;">
                        <div style="font-size:16px;">📆</div><div style="font-size:11px;font-weight:700;color:#a78bfa;margin-top:4px;">Date-wise</div>
                        <div style="font-size:10px;color:#64748b;margin-top:2px;">Specific calendar date</div>
                    </div>
                    <div style="background:rgba(251,146,60,0.08);border:1px solid rgba(251,146,60,0.2);border-radius:10px;padding:10px;text-align:center;">
                        <div style="font-size:16px;">🕐</div><div style="font-size:11px;font-weight:700;color:#fb923c;margin-top:4px;">Period-wise</div>
                        <div style="font-size:10px;color:#64748b;margin-top:2px;">One subject per period</div>
                    </div>
                    <div style="background:rgba(52,211,153,0.08);border:1px solid rgba(52,211,153,0.2);border-radius:10px;padding:10px;text-align:center;">
                        <div style="font-size:16px;">🗓️</div><div style="font-size:11px;font-weight:700;color:#34d399;margin-top:4px;">Week-wise</div>
                        <div style="font-size:10px;color:#64748b;margin-top:2px;">Full Mon–Sat week</div>
                    </div>
                    <div style="background:rgba(244,114,182,0.08);border:1px solid rgba(244,114,182,0.2);border-radius:10px;padding:10px;text-align:center;">
                        <div style="font-size:16px;">📊</div><div style="font-size:11px;font-weight:700;color:#f472b6;margin-top:4px;">Month-wise</div>
                        <div style="font-size:10px;color:#64748b;margin-top:2px;">Complete calendar month</div>
                    </div>
                </div>
            </div>
        """, unsafe_allow_html=True)

        scope_tabs = st.tabs(["📅 Day-wise", "📆 Date-wise", "🕐 Period-wise", "🗓️ Week-wise", "📊 Month-wise"])

        # ── 1. DAY-WISE: all recorded dates for a given weekday ──
        with scope_tabs[0]:
            st.markdown("""<div class="animated-card" style="padding:18px;margin-bottom:14px;">
                <div style="font-size:14px;font-weight:700;color:#4f9eff;margin-bottom:4px;">📅 Day-wise Report</div>
                <div style="font-size:12px;color:#64748b;">Shows all subjects &amp; periods recorded on every <b>Monday / Tuesday /…</b> etc. across all dates. Useful for recurring weekly analysis.</div>
            </div>""", unsafe_allow_html=True)
            dw_day = st.selectbox("Select Weekday", ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"], key="dw_day")
            # Show which dates have data for this day
            matching_dates = sorted(set(r.get("date","") for r in st.session_state.period_attendance if r.get("day")==dw_day and r.get("date")))
            if matching_dates:
                st.markdown(f'<div style="color:#64748b;font-size:12px;margin-bottom:10px;">Data found for: <b style="color:#4f9eff;">{", ".join(matching_dates)}</b></div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div style="color:#64748b;font-size:12px;margin-bottom:10px;">No attendance data recorded for any {dw_day} yet — run AI Live Attendance during a scheduled {dw_day} period.</div>', unsafe_allow_html=True)
            if st.button("📊 Generate Day-wise Report", type="primary", use_container_width=True, key="gen_dw_btn"):
                _generate_and_show_report("Particular Day", dw_day, None, _live_tt)

        # ── 2. DATE-WISE: all subjects on a specific calendar date ──
        with scope_tabs[1]:
            st.markdown("""<div class="animated-card" style="padding:18px;margin-bottom:14px;">
                <div style="font-size:14px;font-weight:700;color:#a78bfa;margin-bottom:4px;">📆 Date-wise Report</div>
                <div style="font-size:12px;color:#64748b;">Shows all subjects &amp; periods for a <b>specific calendar date</b> — exactly what the AI logged that day.</div>
            </div>""", unsafe_allow_html=True)
            dt_date = st.date_input("Select Date", key="dt_date")
            dt_scope = dt_date.strftime("%Y-%m-%d")
            dt_day = dt_date.strftime("%A")
            # Show scheduled periods for that date
            dt_periods = _live_tt.get(dt_scope) or _live_tt.get(dt_day, [])
            if dt_periods:
                period_labels = [f"{p.get('subject','')} [{p.get('period','')}] {p.get('start_time','')}–{p.get('end_time','')}" for p in dt_periods if p.get("type","lecture").lower() in ["lecture","lab","practical"]]
                if period_labels:
                    st.markdown(f'<div style="color:#64748b;font-size:12px;margin-bottom:10px;">Scheduled periods: <b style="color:#a78bfa;">{" | ".join(period_labels)}</b></div>', unsafe_allow_html=True)
            if st.button("📊 Generate Date Report", type="primary", use_container_width=True, key="gen_dt_btn"):
                _generate_and_show_report("Particular Date", dt_scope, None, _live_tt)

        # ── 3. PERIOD-WISE: single subject+period on a specific date ──
        with scope_tabs[2]:
            st.markdown("""<div class="animated-card" style="padding:18px;margin-bottom:14px;">
                <div style="font-size:14px;font-weight:700;color:#fb923c;margin-bottom:4px;">🕐 Period-wise Report</div>
                <div style="font-size:12px;color:#64748b;">Shows attendance for a <b>single subject period</b> on a chosen date — drill down to exactly who was Present / Late / Absent for one class.</div>
            </div>""", unsafe_allow_html=True)
            pw_date = st.date_input("Select Date", key="pw_date")
            pw_scope = pw_date.strftime("%Y-%m-%d")
            pw_day = pw_date.strftime("%A")
            pw_periods = _live_tt.get(pw_scope) or _live_tt.get(pw_day, [])
            pw_class_periods = [p for p in pw_periods if p.get("type","lecture").lower() in ["lecture","lab","practical"]]
            if pw_class_periods:
                pw_options = {f"{p.get('subject','')} [{p.get('period','')}] {p.get('start_time','')}–{p.get('end_time','')}": p.get("period","") for p in pw_class_periods}
                pw_sel_label = st.selectbox("Select Period", list(pw_options.keys()), key="pw_period_sel")
                pw_period_id = pw_options[pw_sel_label]
                if st.button("📊 Generate Period Report", type="primary", use_container_width=True, key="gen_pw_btn"):
                    _generate_and_show_report("Particular Period", pw_scope, pw_period_id, _live_tt)
            else:
                st.markdown(f'<div style="background:rgba(251,146,60,0.08);border:1px solid rgba(251,146,60,0.2);border-radius:10px;padding:14px;color:#fb923c;font-size:13px;">No scheduled periods found for {pw_day} ({pw_scope}). Add periods in the Timetable section first.</div>', unsafe_allow_html=True)

        # ── 4. WEEK-WISE: full Mon–Sat for any selected week ──
        with scope_tabs[3]:
            st.markdown("""<div class="animated-card" style="padding:18px;margin-bottom:14px;">
                <div style="font-size:14px;font-weight:700;color:#34d399;margin-bottom:4px;">🗓️ Week-wise Report</div>
                <div style="font-size:12px;color:#64748b;">All subjects &amp; periods across <b>an entire week (Mon–Sat)</b> — shows each day's columns side by side with entry times.</div>
            </div>""", unsafe_allow_html=True)
            wk_date = st.date_input("Pick any day in the target week", key="wk_date")
            wk_scope = wk_date.strftime("%Y-%m-%d")
            ref_wk = datetime.strptime(wk_scope, "%Y-%m-%d")
            wk_s = ref_wk - timedelta(days=ref_wk.weekday())
            wk_e = wk_s + timedelta(days=5)
            wk_days = [(wk_s + timedelta(days=i)) for i in range(6)]
            day_info = []
            for d in wk_days:
                d_str = d.strftime("%Y-%m-%d"); d_day = d.strftime("%A")
                periods = _live_tt.get(d_str) or _live_tt.get(d_day, [])
                class_p = [p for p in periods if p.get("type","lecture").lower() in ["lecture","lab","practical"]]
                has_data = any(r.get("date")==d_str for r in st.session_state.period_attendance)
                day_info.append({"date":d_str,"day":d_day[:3],"periods":len(class_p),"has_data":has_data})
            st.markdown(f'<div style="color:#64748b;font-size:12px;margin-bottom:10px;">Week range: <b style="color:#34d399;">{wk_s.strftime("%d %b")} – {wk_e.strftime("%d %b %Y")}</b></div>', unsafe_allow_html=True)
            wk_cols = st.columns(6)
            for i, di in enumerate(day_info):
                with wk_cols[i]:
                    dot = "🟢" if di["has_data"] else "⚫"
                    st.markdown(f'<div style="text-align:center;padding:8px;background:rgba(255,255,255,0.03);border-radius:8px;"><div style="font-size:11px;color:#64748b;">{di["day"]}</div><div style="font-size:12px;font-weight:700;color:#94a3b8;">{di["date"][8:]}</div><div style="font-size:10px;color:#64748b;">{di["periods"]}p {dot}</div></div>', unsafe_allow_html=True)
            st.markdown("")
            if st.button("📊 Generate Week Report", type="primary", use_container_width=True, key="gen_wk_btn"):
                _generate_and_show_report("Complete Week", wk_scope, None, _live_tt)

        # ── 5. MONTH-WISE: all days in a calendar month ──
        with scope_tabs[4]:
            st.markdown("""<div class="animated-card" style="padding:18px;margin-bottom:14px;">
                <div style="font-size:14px;font-weight:700;color:#f472b6;margin-bottom:4px;">📊 Month-wise Report</div>
                <div style="font-size:12px;color:#64748b;">Complete month summary — all subjects, all days, subject-wise attendance percentage, and defaulter list for the whole month.</div>
            </div>""", unsafe_allow_html=True)
            months_map = {"January":"01","February":"02","March":"03","April":"04","May":"05","June":"06",
                          "July":"07","August":"08","September":"09","October":"10","November":"11","December":"12"}
            mc1, mc2 = st.columns(2)
            with mc1:
                sel_month = st.selectbox("Month", list(months_map.keys()), index=datetime.now().month-1, key="m_month")
            with mc2:
                sel_year = st.number_input("Year", min_value=2020, max_value=2030, value=datetime.now().year, step=1, key="m_year")
            m_scope = f"{int(sel_year)}-{months_map[sel_month]}"
            # Count recorded days in this month
            recorded_days = sorted(set(r.get("date","") for r in st.session_state.period_attendance if r.get("date","").startswith(m_scope)))
            if recorded_days:
                st.markdown(f'<div style="color:#64748b;font-size:12px;margin-bottom:10px;">AI attendance recorded on <b style="color:#f472b6;">{len(recorded_days)} day(s)</b>: {", ".join(recorded_days)}</div>', unsafe_allow_html=True)
            else:
                st.markdown(f'<div style="color:#64748b;font-size:12px;margin-bottom:10px;">No AI attendance data for {sel_month} {int(sel_year)} yet.</div>', unsafe_allow_html=True)
            if st.button("📊 Generate Month Report", type="primary", use_container_width=True, key="gen_mo_btn"):
                _generate_and_show_report("Complete Month", m_scope, None, _live_tt)


def _generate_and_show_report(scope_type: str, scope_value: str, selected_period_id, live_tt: dict):
    """Shared report generation + display + download logic."""
    import io, calendar as cal_mod

    # ── Load live data sources ──
    try:
        with open("data/erp_attendance.json", "r") as _ef:
            erp_data = json.load(_ef)
    except Exception:
        erp_data = {}

    # Always reload period_attendance from data.json to get latest AI records
    try:
        with open("data.json", "r") as _df:
            _fresh = json.load(_df)
        period_attendance = _fresh.get("period_attendance", st.session_state.period_attendance)
    except Exception:
        period_attendance = st.session_state.period_attendance

    # ── Determine date range for scope ──
    scope_dates = []
    if scope_type == "Particular Day":
        scope_dates = sorted(set(r.get("date","") for r in period_attendance if r.get("day")==scope_value and r.get("date")))
    elif scope_type in ["Particular Date", "Particular Period"]:
        scope_dates = [scope_value]
    elif scope_type == "Complete Week":
        ref = datetime.strptime(scope_value, "%Y-%m-%d")
        wk_s = ref - timedelta(days=ref.weekday())
        scope_dates = [(wk_s + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(6)]
    elif scope_type == "Complete Month":
        yr, mo = int(scope_value[:4]), int(scope_value[5:7])
        _, nd = cal_mod.monthrange(yr, mo)
        scope_dates = [datetime(yr, mo, d).strftime("%Y-%m-%d") for d in range(1, nd+1)]

    # ── Collect all period slots with time info ──
    period_slots = []
    seen_slots = set()
    for d_str in scope_dates:
        d_dt = datetime.strptime(d_str, "%Y-%m-%d")
        d_day = d_dt.strftime("%A")
        periods = live_tt.get(d_str) or live_tt.get(d_day, [])
        for p in periods:
            if p.get("type","lecture").lower() not in ["lecture","lab","practical"]:
                continue
            subj = p.get("subject",""); per_id = p.get("period","")
            if not subj or not per_id:
                continue
            slot_key = (d_str, per_id, subj)
            if slot_key not in seen_slots:
                seen_slots.add(slot_key)
                period_slots.append({"date":d_str,"day":d_day,"period":per_id,"subject":subj,
                                      "start":p.get("start_time",""),"end":p.get("end_time","")})
    # Also add any ERP-recorded periods not in timetable
    for d_str in scope_dates:
        d_day = datetime.strptime(d_str,"%Y-%m-%d").strftime("%A")
        for subj, pdict in erp_data.get(d_str,{}).items():
            for per_id in pdict:
                slot_key = (d_str, per_id, subj)
                if slot_key not in seen_slots:
                    seen_slots.add(slot_key)
                    period_slots.append({"date":d_str,"day":d_day,"period":per_id,"subject":subj,"start":"","end":""})

    if not period_slots:
        st.markdown("""<div style="background:rgba(251,146,60,0.08);border:1px solid rgba(251,146,60,0.3);border-radius:12px;padding:20px;text-align:center;margin-top:16px;">
            <div style="color:#fb923c;font-weight:700;">⚠️ No scheduled classes found for this scope.</div>
            <div style="color:#64748b;font-size:12px;margin-top:6px;">Configure timetable and run AI attendance first.</div>
        </div>""", unsafe_allow_html=True)
        return

    period_slots.sort(key=lambda x: (x["date"], x.get("start","")))

    def col_label(slot):
        t = f" {slot['start']}-{slot['end']}" if slot["start"] else ""
        include_date = scope_type in ["Particular Day", "Complete Week", "Complete Month"]
        d = f" ({slot['date']})" if include_date else ""
        return f"{slot['subject']} [{slot['period']}]{t}{d}"

    _raw_labels = [col_label(s) for s in period_slots]
    _seen_lbl: dict = {}
    col_labels = []
    for _lbl in _raw_labels:
        if _lbl in _seen_lbl:
            _seen_lbl[_lbl] += 1
            col_labels.append(f"{_lbl} #{_seen_lbl[_lbl]}")
        else:
            _seen_lbl[_lbl] = 1
            col_labels.append(_lbl)

    # ── Build rows (Present/Late/Absent per slot) ──
    rows_disp, rows_xl = [], []
    for student in st.session_state.students:
        roll_no = str(student.get("roll_no",""))
        base = {"Name":student.get("name",""),"Roll No":roll_no,"PRN":str(student.get("prn","")),"Branch":student.get("branch",""),"Class":student.get("class","")}
        xl = dict(base)
        total_present = 0

        for slot, lbl in zip(period_slots, col_labels):
            erp_rec = erp_data.get(slot["date"],{}).get(slot["subject"],{}).get(slot["period"],{}).get(roll_no)
            if erp_rec:
                raw = erp_rec.get("status","Absent")
                status = "Present" if raw in ["Present","Auto-Carried"] else ("Late" if raw=="Late" else "Absent")
                entry = erp_rec.get("entry_time","-") or "-"
            else:
                status = "Absent"; entry = "-"
                for pa in period_attendance:
                    if pa.get("date")==slot["date"] and pa.get("period")==slot["period"] and pa.get("subject")==slot["subject"]:
                        raw = pa.get("attendance",{}).get(roll_no,"Absent")
                        status = "Present" if raw in ["Present","Auto-Carried"] else ("Late" if raw=="Late" else "Absent")
                        break

            if status in ["Present","Late"]: total_present += 1
            base[lbl] = "✅ Present" if status=="Present" else ("⏱️ Late" if status=="Late" else "❌ Absent")
            xl[lbl] = status
            xl[f"Entry [{slot['period']}]"] = entry

        pct = round(total_present/len(period_slots)*100,1) if period_slots else 100.0
        base["Att %"] = f"{pct}%"; base["Status"] = "OK ✅" if pct>=75 else "Defaulter ⚠️"
        xl["Attendance %"] = pct; xl["Status"] = "OK" if pct>=75 else "Defaulter"
        rows_disp.append(base); rows_xl.append(xl)

    df_disp = pd.DataFrame(rows_disp)
    df_xl = pd.DataFrame(rows_xl)
    df_def_disp = df_disp[df_disp["Status"].str.startswith("Defaulter")]
    df_def_xl = df_xl[df_xl["Status"]=="Defaulter"]

    total_s = len(rows_disp); defaulters = len(df_def_disp)
    avg_pct = round(df_xl["Attendance %"].mean(),1) if not df_xl.empty else 0

    st.markdown("---")
    k1,k2,k3,k4 = st.columns(4)
    with k1: st.metric("👥 Students", total_s)
    with k2: st.metric("✅ OK (≥75%)", total_s-defaulters)
    with k3: st.metric("⚠️ Defaulters", defaulters)
    with k4: st.metric("📊 Avg Att.", f"{avg_pct}%")
    st.markdown("---")

    # ── Full report HTML table ──
    _info_cols = ["Name","Roll No","PRN","Branch","Class"]
    _period_lbls = [c for c in df_disp.columns if c not in _info_cols + ["Att %","Status"]]

    def _build_report_table(df_src, cols, title, accent):
        hdr = "".join(f'<th style="padding:9px 12px;font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.07em;color:#64748b;white-space:nowrap;border-bottom:2px solid rgba(255,255,255,0.06);">{c}</th>' for c in ["Name","Roll No"])
        for c in _period_lbls:
            hdr += f'<th style="padding:9px 12px;font-size:10px;font-weight:700;color:{accent};white-space:nowrap;border-bottom:2px solid rgba(79,158,255,0.18);">{c}</th>'
        hdr += '<th style="padding:9px 12px;font-size:10px;font-weight:700;color:#94a3b8;white-space:nowrap;border-bottom:2px solid rgba(255,255,255,0.06);">Att %</th>'
        hdr += '<th style="padding:9px 12px;font-size:10px;font-weight:700;color:#94a3b8;white-space:nowrap;border-bottom:2px solid rgba(255,255,255,0.06);">Status</th>'
        rows_html = ""
        for i, (_, r) in enumerate(df_src.iterrows()):
            bg = "rgba(255,255,255,0.015)" if i % 2 == 0 else "transparent"
            cells = f'<td style="padding:9px 12px;color:#e2e8f0;font-size:12px;white-space:nowrap;">{r["Name"]}</td>'
            cells += f'<td style="padding:9px 12px;color:#64748b;font-size:12px;">{r["Roll No"]}</td>'
            for c in _period_lbls:
                v = r.get(c, "❌ Absent")
                if "Present" in str(v):
                    badge = f'<span style="background:rgba(52,211,153,0.12);color:#34d399;border:1px solid rgba(52,211,153,0.25);border-radius:5px;padding:2px 8px;font-size:10px;font-weight:700;">✅ P</span>'
                elif "Late" in str(v):
                    badge = f'<span style="background:rgba(251,146,60,0.12);color:#fb923c;border:1px solid rgba(251,146,60,0.25);border-radius:5px;padding:2px 8px;font-size:10px;font-weight:700;">⏱ L</span>'
                else:
                    badge = f'<span style="background:rgba(244,114,182,0.08);color:#f472b6;border:1px solid rgba(244,114,182,0.18);border-radius:5px;padding:2px 8px;font-size:10px;font-weight:700;">❌ A</span>'
                cells += f'<td style="padding:6px 12px;text-align:center;">{badge}</td>'
            att = r.get("Att %", "")
            sta = r.get("Status", "")
            sta_color = "#34d399" if "OK" in str(sta) else "#f472b6"
            cells += f'<td style="padding:9px 12px;color:#94a3b8;font-size:12px;text-align:center;">{att}</td>'
            cells += f'<td style="padding:9px 12px;font-size:11px;font-weight:700;color:{sta_color};text-align:center;">{sta}</td>'
            rows_html += f'<tr style="background:{bg};border-bottom:1px solid rgba(255,255,255,0.03);">{cells}</tr>'
        return f"""
        <div style="background:rgba(12,17,33,0.85);border:1px solid rgba(79,158,255,0.1);border-radius:14px;overflow:hidden;margin-bottom:12px;">
            <div style="padding:12px 16px;border-bottom:1px solid rgba(79,158,255,0.08);">
                <span style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:{accent};">{title}</span>
            </div>
            <div style="overflow-x:auto;"><table style="width:100%;border-collapse:collapse;font-family:Inter,sans-serif;">
                <thead><tr style="background:rgba(6,9,18,0.5);">{hdr}</tr></thead>
                <tbody>{rows_html}</tbody>
            </table></div>
        </div>"""

    st.markdown(_build_report_table(df_disp, df_disp.columns, "📋 Complete Attendance Sheet", "#4f9eff"), unsafe_allow_html=True)

    if not df_def_disp.empty:
        st.markdown(_build_report_table(df_def_disp, df_def_disp.columns, "⚠️ Defaulters — Below 75%", "#f472b6"), unsafe_allow_html=True)
    else:
        st.markdown("""<div style="background:rgba(52,211,153,0.06);border:1px solid rgba(52,211,153,0.18);border-radius:12px;padding:18px;text-align:center;margin-bottom:12px;">
            <div style="font-size:22px;">🎉</div><div style="color:#34d399;font-weight:700;font-size:13px;">No Defaulters!</div>
            <div style="color:#64748b;font-size:11px;margin-top:4px;">All students have ≥75% attendance</div></div>""", unsafe_allow_html=True)

    st.markdown("---")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df_xl.to_excel(w, sheet_name='Attendance_Report', index=False)
        pd.DataFrame(period_slots).to_excel(w, sheet_name='Period_Schedule', index=False)
        if not df_def_xl.empty:
            df_def_xl[["Name","Roll No","PRN","Branch","Class","Attendance %","Status"]].to_excel(w, sheet_name='Defaulters', index=False)
    buf.seek(0)

    fn_tag = scope_value.replace("-","_").replace(" ","_").lower()
    stype_tag = scope_type.lower().replace(" ","_")
    st.markdown('<p style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#64748b;margin-bottom:10px;">📥 Download Excel</p>', unsafe_allow_html=True)
    st.download_button("📥 Download Full Attendance Report (.xlsx)", data=buf,
        file_name=f"attendance_{stype_tag}_{fn_tag}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True, key=f"dl_full_{stype_tag}_{fn_tag}")


def show_ai_live_attendance_page():
    """AI Live Attendance — Cloud-compatible: uses browser camera via st.camera_input()."""

    st.markdown("""
        <style>
            .compact-card {
                background: rgba(12, 17, 33, 0.75);
                border: 1px solid rgba(79, 158, 255, 0.15);
                border-radius: 16px; padding: 16px; margin-bottom: 12px;
                backdrop-filter: blur(12px); box-shadow: 0 4px 24px rgba(0,0,0,0.3);
            }
            .section-label {
                font-size: 11px; font-weight: 700; text-transform: uppercase;
                letter-spacing: 0.1em; color: #64748b; margin-bottom: 8px;
            }
        </style>
    """, unsafe_allow_html=True)

    st.markdown("""
        <div style="margin-bottom:24px;">
            <h2 style="margin:0;font-size:28px;font-weight:900;background:linear-gradient(135deg,#34d399,#4f9eff);-webkit-background-clip:text;-webkit-text-fill-color:transparent;">
                🤖 AI Live Attendance
            </h2>
            <p style="color:#64748b;font-size:13px;margin:4px 0 0 0;">
                Real-time face recognition via browser camera — works on any device through a link
            </p>
        </div>
    """, unsafe_allow_html=True)

    # ── Check prerequisites ──
    if len(st.session_state.students) == 0:
        st.warning("⚠️ No students registered. Add students first.")
        return
    if not Path("data/face_embeddings.pkl").exists():
        st.error("❌ No face embeddings found. Add students and capture photos to train the AI.")
        return

    # ── Load known embeddings (cached in session state) ──
    if "known_faces" not in st.session_state or st.session_state.get("reload_embeddings"):
        try:
            from vision.cloud_recognition import load_embeddings
            st.session_state.known_faces = load_embeddings()
            st.session_state.reload_embeddings = False
        except Exception as e:
            st.error(f"Failed to load face embeddings: {e}")
            return

    known_faces = st.session_state.known_faces
    if not known_faces:
        st.warning("⚠️ No student face data found. Please capture photos and generate embeddings first.")
        return

    # ── Check current timetable period ──
    try:
        from logic.attendance_engine import AttendanceLogicEngine
        if "att_engine" not in st.session_state:
            st.session_state.att_engine = AttendanceLogicEngine()
        att_engine = st.session_state.att_engine
        att_engine._load_timetable_from_app()
        current_period = att_engine.get_current_period()
    except Exception as e:
        st.warning(f"Attendance engine error: {e}")
        att_engine = None
        current_period = None

    # ── Layout ──
    left_col, right_col = st.columns([1.1, 0.9])

    with left_col:
        st.markdown('<p class="section-label">📷 Browser Camera Feed</p>', unsafe_allow_html=True)

        # Browser camera input — works on desktop & mobile via browser link
        cam_frame = st.camera_input(
            "📸 Point camera at student face",
            key="live_cam_input",
            help="Allow camera access in your browser when prompted"
        )

        # Period info banner
        if current_period:
            subj = current_period.get("subject", "—")
            per_id = current_period.get("period", "—")
            start = current_period.get("start_time", "")
            end = current_period.get("end_time", "")
            st.markdown(f"""
                <div style="background:rgba(79,158,255,0.08);border:1px solid rgba(79,158,255,0.2);
                border-radius:10px;padding:12px 16px;margin-top:10px;">
                    <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:0.08em;">Active Period</div>
                    <div style="font-size:16px;font-weight:800;color:#4f9eff;margin-top:2px;">{subj}</div>
                    <div style="font-size:12px;color:#64748b;">{per_id} · {start}–{end}</div>
                </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
                <div style="background:rgba(251,146,60,0.06);border:1px solid rgba(251,146,60,0.2);
                border-radius:10px;padding:12px 16px;margin-top:10px;">
                    <div style="font-size:12px;color:#fb923c;font-weight:700;">⚠️ No Active Period</div>
                    <div style="font-size:11px;color:#64748b;margin-top:2px;">Check timetable — attendance is only logged during scheduled periods.</div>
                </div>
            """, unsafe_allow_html=True)

    with right_col:
        st.markdown('<p class="section-label">📊 System Status</p>', unsafe_allow_html=True)

        # Live metrics
        active_count = len(att_engine.active_students) if att_engine else 0
        logged_today = 0
        if att_engine:
            today_erp = att_engine.erp_data.get(datetime.now().strftime("%Y-%m-%d"), {})
            logged_today = sum(len(s) for p in today_erp.values() for s in p.values())

        mc1, mc2 = st.columns(2)
        with mc1:
            st.metric("Active Students", active_count)
            st.metric("Attendance Logged", logged_today)
        with mc2:
            st.metric("Known Faces", len(known_faces))
            period_label = current_period.get("period", "None") if current_period else "None"
            st.metric("Current Period", period_label)

        st.markdown('<p class="section-label" style="margin-top:16px;">📝 Live Events</p>', unsafe_allow_html=True)

        if att_engine:
            events = att_engine.event_logs
            if events:
                for ev in reversed(events[-6:]):
                    e_type = ev.get("type", "")
                    nm = ev.get("name", "")
                    subj_ev = ev.get("subject", "")
                    per_ev = ev.get("period", "")
                    t_ev = ev.get("time", "")
                    if e_type == "ENTRY":
                        color, icon = "#34d399", "✅ PRESENT"
                    elif e_type == "LATE":
                        color, icon = "#fb923c", "⏱️ LATE"
                    elif e_type == "EXIT":
                        color, icon = "#64748b", "🚶 EXIT"
                    else:
                        color, icon = "#94a3b8", e_type
                    st.markdown(f"""
                        <div style="background:rgba(255,255,255,0.03);border-left:3px solid {color};
                        border-radius:6px;padding:7px 12px;margin-bottom:5px;">
                            <span style="color:{color};font-weight:700;font-size:11px;">{icon}</span>
                            <span style="color:#ffffff;font-size:12px;"> {nm}</span>
                            <span style="color:#64748b;font-size:10px;"> · {subj_ev} [{per_ev}] {t_ev}</span>
                        </div>
                    """, unsafe_allow_html=True)
            else:
                st.markdown('<div style="color:#64748b;font-size:12px;padding:8px;">No events yet. Point camera at student during a scheduled period.</div>', unsafe_allow_html=True)

    # ── Process captured frame ──
    if cam_frame is not None:
        st.markdown("---")
        st.markdown('<p class="section-label">🔍 Recognition Result</p>', unsafe_allow_html=True)

        res_col1, res_col2 = st.columns([1, 1])

        with st.spinner("🧠 Processing face recognition..."):
            try:
                from vision.cloud_recognition import process_camera_frame
                result = process_camera_frame(cam_frame, known_faces)
            except Exception as e:
                st.error(f"Recognition error: {e}")
                result = {"match": None, "annotated_frame": None, "face_count": 0}

        with res_col1:
            if result["annotated_frame"] is not None:
                st.image(result["annotated_frame"], caption=f"Detected {result['face_count']} face(s)", use_container_width=True)

        with res_col2:
            match = result["match"]
            if match:
                roll_no = match["roll_no"]
                name = match["name"]
                sim = match["similarity"]

                # Look up full student info
                student_info = next((s for s in st.session_state.students if str(s.get("roll_no")) == str(roll_no)), {})

                # Show student card
                st.markdown(f"""
                    <div style="background:rgba(52,211,153,0.08);border:1px solid rgba(52,211,153,0.25);
                    border-radius:12px;padding:16px;margin-bottom:12px;">
                        <div style="color:#34d399;font-weight:800;font-size:16px;">✅ Recognized!</div>
                        <div style="color:#ffffff;font-size:18px;font-weight:700;margin-top:6px;">{name}</div>
                        <div style="color:#64748b;font-size:12px;">Roll: {roll_no} · Match: {sim*100:.1f}%</div>
                        <div style="color:#64748b;font-size:12px;margin-top:2px;">
                            Branch: {student_info.get("branch","—")} · Class: {student_info.get("class","—")}
                        </div>
                    </div>
                """, unsafe_allow_html=True)

                # Log attendance if period is active
                if att_engine and current_period:
                    try:
                        att_engine.process_recognition_event(roll_no=roll_no, name=name)
                        subj_now = current_period.get("subject", "")
                        per_now = current_period.get("period", "")
                        erp_now = att_engine.erp_data.get(datetime.now().strftime("%Y-%m-%d"), {})
                        rec = erp_now.get(subj_now, {}).get(per_now, {}).get(roll_no)
                        if rec:
                            status_color = "#34d399" if rec["status"] == "Present" else "#fb923c"
                            st.markdown(f"""
                                <div style="background:rgba(79,158,255,0.06);border:1px solid rgba(79,158,255,0.15);
                                border-radius:10px;padding:10px 14px;">
                                    <div style="font-size:11px;color:#64748b;">Attendance Status</div>
                                    <div style="font-size:15px;font-weight:700;color:{status_color};">
                                        {rec['status']}
                                    </div>
                                    <div style="font-size:11px;color:#64748b;">Logged at {rec.get('entry_time','—')} · {subj_now} [{per_now}]</div>
                                </div>
                            """, unsafe_allow_html=True)
                    except Exception as e:
                        st.warning(f"Attendance logging issue: {e}")
                elif not current_period:
                    st.info("ℹ️ Student recognized but no active period — attendance not logged.")
            elif result["face_count"] == 0:
                st.markdown("""
                    <div style="background:rgba(100,116,139,0.08);border:1px solid rgba(100,116,139,0.2);
                    border-radius:12px;padding:16px;text-align:center;">
                        <div style="font-size:24px;">👤</div>
                        <div style="color:#94a3b8;font-weight:700;margin-top:6px;">No Face Detected</div>
                        <div style="color:#64748b;font-size:12px;margin-top:4px;">Make sure your face is clearly visible and well-lit.</div>
                    </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("""
                    <div style="background:rgba(244,114,182,0.06);border:1px solid rgba(244,114,182,0.2);
                    border-radius:12px;padding:16px;text-align:center;">
                        <div style="font-size:24px;">❓</div>
                        <div style="color:#f472b6;font-weight:700;margin-top:6px;">Unknown Person</div>
                        <div style="color:#64748b;font-size:12px;margin-top:4px;">Face detected but not in the student database.</div>
                    </div>
                """, unsafe_allow_html=True)




def show_ai_add_student_page():
    """Display the AI Student Dataset Collection page."""
    st.title("🤖 AI Add Student (Dataset Collector)")
    st.markdown("Use this panel to capture high-quality face images for the AI recognition engine.")
    
    with st.form("add_student_form"):
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            name = st.text_input("Full Name*", placeholder="e.g. John Doe")
        with col2:
            roll_no = st.text_input("Roll Number*", placeholder="e.g. 101")
        with col3:
            prn = st.text_input("College PRN", placeholder="e.g. PRN12345")
        with col4:
            student_class = st.selectbox("Class", ["FY", "SY", "TY", "B.Tech", "FY B.Tech", "SY B.Tech", "TY B.Tech"])
            
        st.markdown("### 📷 Capture Face Image")
        st.info("Please look directly at the camera. Ensure good lighting and a clear background.")
        
        # Streamlit native camera input
        camera_photo = st.camera_input("Take Student Photo")
        
        submit = st.form_submit_button("Save Student Data", type="primary", use_container_width=True)
        
        if submit:
            if not name or not roll_no:
                st.error("Name and Roll Number are required fields.")
            elif camera_photo is None:
                st.error("Please capture a photo first.")
            else:
                try:
                    import cv2
                    import numpy as np
                    from PIL import Image
                    
                    # Convert uploaded image to numpy array
                    image = Image.open(camera_photo)
                    img_array = np.array(image)
                    
                    # Convert RGB to BGR for OpenCV
                    img_bgr = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
                    
                    # Resize to 1000x1000 as requested
                    resized_img = cv2.resize(img_bgr, (1000, 1000), interpolation=cv2.INTER_LINEAR)
                    
                    # Create directory
                    safe_name = "".join([c for c in name if c.isalpha() or c.isdigit() or c==' ']).rstrip()
                    dir_name = f"{roll_no}_{safe_name.replace(' ', '_')}"
                    save_dir = os.path.join("dataset", dir_name)
                    os.makedirs(save_dir, exist_ok=True)
                    
                    # Generate filename
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"face_{roll_no}_{timestamp}.jpg"
                    filepath = os.path.join(save_dir, filename)
                    
                    # Save image
                    cv2.imwrite(filepath, resized_img, [cv2.IMWRITE_JPEG_QUALITY, 100])
                    
                    # Success message
                    st.success(f"✅ Successfully added student: {name}")
                    st.info(f"Image saved at 1000x1000px resolution in `{filepath}`.")
                    
                    # Also save metadata if PRN is provided
                    meta_path = os.path.join(save_dir, "metadata.json")
                    with open(meta_path, 'w') as f:
                        json.dump({"name": name, "roll_no": roll_no, "prn": prn, "class": student_class, "added_on": timestamp}, f)
                    
                    st.info("🔄 Generating AI Face Embeddings... This may take a moment.")
                    from vision.generate_embeddings import generate_face_embeddings
                    result = generate_face_embeddings(verbose=False)
                    
                    if result.get('status') in ['success', 'partial_success']:
                        st.success("✅ AI Embeddings updated successfully! The student will now be recognized.")
                        # Force a reset of the engine so it reloads the new embeddings
                        st.session_state.recognition_engine = None
                    else:
                        st.error(f"⚠️ Warning: Failed to generate embeddings: {result.get('error', 'Unknown Error')}")
                        
                except Exception as e:
                    st.error(f"Failed to process image: {e}")



# ============================================================================
# J. PAGE ROUTING & MAIN APP
# ============================================================================

def main():
    """Main application entry point with horizontal page routing."""
    
    # 1. High-Fidelity Futuristic Header & Live Status Indicator
    st.markdown("""
        <div style="display: flex; align-items: center; justify-content: space-between; margin-bottom: 25px; border-bottom: 1px solid rgba(255,255,255,0.04); padding-bottom: 15px;">
            <div style="display: flex; align-items: center; gap: 15px;">
                <span style="font-size: 38px; filter: drop-shadow(0 0 10px rgba(167, 139, 250, 0.4));">🤖</span>
                <h1 style="margin: 0; font-size: 34px; font-weight: 900; letter-spacing: -0.04em; background: linear-gradient(135deg, #4f9eff 0%, #a78bfa 100%); -webkit-background-clip: text; -webkit-text-fill-color: transparent;">AttendAI</h1>
            </div>
            <div style="display: flex; align-items: center; gap: 10px; background: rgba(255,255,255,0.03); border: 1px solid rgba(255,255,255,0.05); border-radius: 30px; padding: 6px 16px;">
                <span class="active-pulse" style="display: inline-block; width: 8px; height: 8px; background: #34d399; border-radius: 50%; box-shadow: 0 0 8px #34d399; animation: pulse 1.5s infinite alternate;"></span>
                <span style="font-size: 11px; font-weight: 700; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.08em;">AI Neural Grid Active</span>
            </div>
        </div>
        <style>
            @keyframes pulse {
                0% { opacity: 0.4; }
                100% { opacity: 1; }
            }
        </style>
    """, unsafe_allow_html=True)
    
    # 2. Render horizontal navigation
    page_options = [
        "Dashboard",
        "Manage Students",
        "Timetable",
        "🤖 AI Live Attendance",
        "📊 Reports"
    ]
    
    current_page = st.session_state.get('page', 'Dashboard')
    
    col1, col2, col3, col4, col5 = st.columns(5)
    icons = ["📊", "👥", "📅", "🤖", "📈"]
    
    for idx, (p, ic) in enumerate(zip(page_options, icons)):
        with [col1, col2, col3, col4, col5][idx]:
            is_active = (current_page == p)
            btn_style = "primary" if is_active else "secondary"
            if st.button(f"{ic} {p}", type=btn_style, use_container_width=True, key=f"nav_{p}"):
                st.session_state.page = p
                st.rerun()
                
    st.markdown("---")
    
    # Page routing map
    page_routing = {
        "Dashboard": show_dashboard,
        "Manage Students": show_manage_students,
        "Timetable": show_timetable,
        "🤖 AI Live Attendance": show_ai_live_attendance_page,
        "📊 Reports": show_reports_page,
    }
    
    page_function = page_routing.get(st.session_state.page, show_dashboard)
    page_function()


# ============================================================================
# K. APPLICATION ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    main()
